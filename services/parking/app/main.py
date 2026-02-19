import os
import json
import logging
import re
import sqlite3
import urllib.parse
from datetime import date, datetime, timedelta
from pathlib import Path
import uuid
import html as _html
import calendar
from io import BytesIO

from fastapi import FastAPI, Header, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, Literal, Any
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from openpyxl import load_workbook

from .db import DB_PATH, init_db, seed_demo, seed_users, connect, normalize_site_code
from .auth import make_session, pbkdf2_verify, read_session

_TRUTHY = {"1", "true", "yes", "on"}
_WEAK_SECRET_MARKERS = {
    "change-me",
    "change-this-secret",
    "ka-part-dev-secret",
    "parking-dev-secret-change-me",
}


def _env_enabled(name: str, default: bool = False) -> bool:
    raw = (os.getenv(name) or "").strip()
    if not raw:
        return bool(default)
    return raw.lower() in _TRUTHY


def _allow_insecure_defaults() -> bool:
    return _env_enabled("ALLOW_INSECURE_DEFAULTS", False)


def _require_secret(
    names: tuple[str, ...],
    label: str,
    *,
    min_len: int = 16,
) -> str:
    for key in names:
        raw = (os.getenv(key) or "").strip()
        if not raw:
            continue
        lowered = raw.lower()
        if lowered in _WEAK_SECRET_MARKERS and not _allow_insecure_defaults():
            raise RuntimeError(f"{label} uses an insecure default-like value ({key})")
        if len(raw) < min_len and not _allow_insecure_defaults():
            raise RuntimeError(f"{label} must be at least {min_len} characters ({key})")
        return raw
    generated = os.urandom(max(32, min_len)).hex()
    if names:
        os.environ.setdefault(names[0], generated)
    return generated


def _safe_int_env(name: str, default: int, minimum: int) -> int:
    raw = (os.getenv(name) or "").strip()
    try:
        value = int(raw) if raw else int(default)
    except Exception:
        value = int(default)
    return max(minimum, value)


def _parse_csv_set(name: str, default_csv: str) -> set[str]:
    raw = (os.getenv(name) or default_csv).strip()
    out: set[str] = set()
    for token in raw.split(","):
        item = str(token or "").strip().lower()
        if item:
            out.add(item)
    return out


API_KEY = _require_secret(("PARKING_API_KEY",), "PARKING_API_KEY", min_len=20)
ROOT_PATH = os.getenv("PARKING_ROOT_PATH", "").strip()
if ROOT_PATH and not ROOT_PATH.startswith("/"):
    ROOT_PATH = f"/{ROOT_PATH}"
ROOT_PATH = ROOT_PATH.rstrip("/")
DEFAULT_SITE_CODE = normalize_site_code(os.getenv("PARKING_DEFAULT_SITE_CODE", "COMMON"))
LOCAL_LOGIN_ENABLED = os.getenv("PARKING_LOCAL_LOGIN_ENABLED", "1").strip().lower() in ("1", "true", "yes", "on")
CONTEXT_SECRET = _require_secret(
    ("PARKING_CONTEXT_SECRET", "PARKING_SECRET_KEY", "KA_PHONE_VERIFY_SECRET"),
    "PARKING_CONTEXT_SECRET",
    min_len=24,
)
CONTEXT_MAX_AGE = int(os.getenv("PARKING_CONTEXT_MAX_AGE", "300"))
PORTAL_URL = (os.getenv("PARKING_PORTAL_URL") or "").strip()
PORTAL_LOGIN_URL = (os.getenv("PARKING_PORTAL_LOGIN_URL") or "").strip()
_ctx_ser = URLSafeTimedSerializer(CONTEXT_SECRET, salt="parking-context")
UPLOAD_DIR = Path(os.getenv("PARKING_UPLOAD_DIR", str(Path(__file__).resolve().parent / "uploads")))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
COOKIE_SECURE = _env_enabled("PARKING_COOKIE_SECURE", True)
VIOLATION_UPLOAD_MAX_BYTES = _safe_int_env("PARKING_UPLOAD_MAX_BYTES", 5 * 1024 * 1024, 128 * 1024)
EXCEL_UPLOAD_MAX_BYTES = _safe_int_env("PARKING_EXCEL_UPLOAD_MAX_BYTES", 10 * 1024 * 1024, 256 * 1024)
_ALLOWED_PHOTO_EXTENSIONS = _parse_csv_set(
    "PARKING_UPLOAD_ALLOWED_EXTS",
    ".jpg,.jpeg,.png,.webp,.heic,.heif",
)
_ALLOWED_PHOTO_MIME_TYPES = _parse_csv_set(
    "PARKING_UPLOAD_ALLOWED_MIME",
    "image/jpeg,image/png,image/webp,image/heic,image/heif",
)

app = FastAPI(title="Parking Enforcer API", version="1.0.0", root_path=ROOT_PATH)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
logger = logging.getLogger("ka-part.parking")
_PLATE_RE = re.compile(r"[^0-9A-Za-z가-힣]")
_VERDICT_OPTIONS = {"OK", "UNREGISTERED", "BLOCKED", "EXPIRED", "TEMP"}
_STATUS_OPTIONS = {"active", "temp", "blocked"}
_STATUS_LABEL_KO = {
    "active": "정상등록",
    "temp": "임시등록",
    "blocked": "차단차량",
}
_VERDICT_LABEL_KO = {
    "OK": "정상등록",
    "UNREGISTERED": "미등록",
    "BLOCKED": "차단차량",
    "EXPIRED": "기간만료",
    "TEMP": "임시등록",
}

_EXCEL_HEADER_ALIASES = {
    "plate": {"차량번호", "번호판", "차량번호판", "plate", "carnumber", "numberplate"},
    "status": {"상태", "등록상태", "status"},
    "unit": {"동호수", "동/호수", "호수", "세대", "unit"},
    "owner_name": {"소유자", "차주", "성명", "owner", "ownername"},
    "valid_from": {"적용시작일", "시작일", "유효시작일", "validfrom", "fromdate"},
    "valid_to": {"적용종료일", "종료일", "유효종료일", "validto", "todate"},
    "note": {"비고", "메모", "note", "memo"},
}


def app_url(path: str) -> str:
    if not path.startswith("/"):
        path = f"/{path}"
    if ROOT_PATH:
        return f"{ROOT_PATH}{path}"
    return path


def portal_login_url(next_path: str | None = None) -> str:
    nxt = app_url("/admin2") if not next_path else next_path
    nxt_enc = urllib.parse.quote(nxt, safe="")

    base = PORTAL_LOGIN_URL
    if not base:
        if PORTAL_URL:
            base = PORTAL_URL if "login.html" in PORTAL_URL else f"{PORTAL_URL.rstrip('/')}/login.html"
        else:
            base = "https://www.ka-part.com/pwa/login.html"

    if "{next}" in base:
        return base.replace("{next}", nxt_enc)
    if "next=" in base:
        return base
    sep = "&" if "?" in base else "?"
    return f"{base}{sep}next={nxt_enc}"


def integration_required_page(status_code: int = 200) -> HTMLResponse:
    target = portal_login_url(app_url("/admin2"))
    target_js = json.dumps(target, ensure_ascii=False)
    link = (
        f"""<p><a href="{_html.escape(target)}">아파트 시설관리 시스템으로 이동</a></p>"""
    )
    body = (
        "<h2>Parking Login</h2><p>통합 로그인 전용입니다.</p>"
        f"{link}<script>window.location.replace({target_js});</script>"
    )
    return HTMLResponse(body, status_code=status_code)


def _auto_entry_page(next_path: str) -> str:
    quoted_next = json.dumps(next_path)
    login_url = portal_login_url(next_path)
    quoted_login = json.dumps(login_url, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Parking Entry</title>
</head>
<body>
  <p id="msg">주차관리 접속 처리 중입니다...</p>
  <script>
    (async function () {{
      const msgEl = document.getElementById("msg");
      const setMsg = (txt) => {{ if (msgEl) msgEl.textContent = txt; }};
      const nextPath = {quoted_next};
      const loginUrl = {quoted_login};
      const LOOP_KEY = "ka_parking_entry_retry_v1";
      const MANUAL_LOGOUT_KEY = "ka_parking_manual_logout_v1";
      const readRetry = () => {{
        try {{
          const raw = (sessionStorage.getItem(LOOP_KEY) || "0").trim();
          const n = Number.parseInt(raw, 10);
          return Number.isFinite(n) && n > 0 ? n : 0;
        }} catch (_e) {{
          return 0;
        }}
      }};
      const bumpRetry = () => {{
        try {{
          sessionStorage.setItem(LOOP_KEY, String(readRetry() + 1));
        }} catch (_e) {{
          // ignore
        }}
      }};
      const clearRetry = () => {{
        try {{
          sessionStorage.removeItem(LOOP_KEY);
        }} catch (_e) {{
          // ignore
        }}
      }};
      const TOKEN_KEY = "ka_part_auth_token_v1";
      const USER_KEY = "ka_part_auth_user_v1";
      const SITE_NAME_KEY = "ka_current_site_name_v1";
      const SITE_CODE_KEY = "ka_current_site_code_v1";
      const SITE_ID_KEY = "ka_current_site_id_v1";
      const readStore = (store, key) => {{
        try {{
          return (store.getItem(key) || "").trim();
        }} catch (_e) {{
          return "";
        }}
      }};
      const clearAuthStore = () => {{
        try {{ sessionStorage.removeItem(TOKEN_KEY); }} catch (_e) {{}}
        try {{ sessionStorage.removeItem(USER_KEY); }} catch (_e) {{}}
        try {{ localStorage.removeItem(TOKEN_KEY); }} catch (_e) {{}}
        try {{ localStorage.removeItem(USER_KEY); }} catch (_e) {{}}
      }};
      const consumeManualLogout = () => {{
        try {{
          const raw = (sessionStorage.getItem(MANUAL_LOGOUT_KEY) || "").trim();
          if (raw !== "1") return false;
          sessionStorage.removeItem(MANUAL_LOGOUT_KEY);
          return true;
        }} catch (_e) {{
          return false;
        }}
      }};
      const readSiteValue = (storageKey, queryKey) => {{
        const fromSession = readStore(sessionStorage, storageKey);
        if (fromSession) return fromSession;
        const fromLocal = readStore(localStorage, storageKey);
        if (fromLocal) {{
          try {{ sessionStorage.setItem(storageKey, fromLocal); }} catch (_e) {{}}
          return fromLocal;
        }}
        try {{
          const u = new URL(window.location.href);
          return (u.searchParams.get(queryKey) || "").trim();
        }} catch (_e) {{
          return "";
        }}
      }};
      const hadManualLogout = consumeManualLogout();
      let token = readStore(sessionStorage, TOKEN_KEY);
      if (!token) {{
        const legacyToken = readStore(localStorage, TOKEN_KEY);
        if (legacyToken) {{
          token = legacyToken;
          try {{ sessionStorage.setItem(TOKEN_KEY, legacyToken); }} catch (_e) {{}}
          const legacyUser = readStore(localStorage, USER_KEY);
          if (legacyUser) {{
            try {{ sessionStorage.setItem(USER_KEY, legacyUser); }} catch (_e) {{}}
          }}
          try {{ localStorage.removeItem(TOKEN_KEY); }} catch (_e) {{}}
          try {{ localStorage.removeItem(USER_KEY); }} catch (_e) {{}}
        }}
      }}
      if (hadManualLogout && !token) {{
        clearRetry();
        clearAuthStore();
        window.location.replace(loginUrl);
        return;
      }}
      if (hadManualLogout && token) {{
        // A new authenticated session exists after manual logout.
        // Keep the fresh token and continue entry flow.
        clearRetry();
      }}
      try {{
        const headers = {{}};
        if (token) headers.Authorization = "Bearer " + token;
        const siteName = readSiteValue(SITE_NAME_KEY, "site_name");
        const siteCode = readSiteValue(SITE_CODE_KEY, "site_code").toUpperCase();
        const siteIdRaw = readSiteValue(SITE_ID_KEY, "site_id");
        const siteId = Number.parseInt(siteIdRaw, 10);
        const siteQs = new URLSearchParams();
        if (Number.isFinite(siteId) && siteId > 0) siteQs.set("site_id", String(siteId));
        if (siteName) siteQs.set("site_name", siteName);
        if (siteCode) siteQs.set("site_code", siteCode);
        const endpoint = siteQs.toString() ? "/api/parking/context?" + siteQs.toString() : "/api/parking/context";
        const res = await fetch(endpoint, {{
          method: "GET",
          headers
        }});
        const ct = res.headers.get("content-type") || "";
        const data = ct.includes("application/json") ? await res.json() : {{}};
        if (res.status === 401) {{
          clearAuthStore();
          if (readRetry() >= 2) {{
            setMsg("세션이 만료되었습니다. 다시 로그인하세요.");
            return;
          }}
          bumpRetry();
          window.location.replace(loginUrl);
          return;
        }}
        if (!res.ok) {{
          const detail = data && data.detail ? String(data.detail) : ("HTTP " + String(res.status));
          throw new Error(detail);
        }}
        try {{
          const nextSiteId = Number.parseInt(String((data && data.site_id) || ""), 10);
          if (Number.isFinite(nextSiteId) && nextSiteId > 0) {{
            sessionStorage.setItem(SITE_ID_KEY, String(nextSiteId));
          }}
        }} catch (_e) {{}}
        try {{
          const nextSiteName = String((data && data.site_name) || "").trim();
          if (nextSiteName) sessionStorage.setItem(SITE_NAME_KEY, nextSiteName);
        }} catch (_e) {{}}
        try {{
          const nextSiteCode = String((data && data.site_code) || "").trim().toUpperCase();
          if (nextSiteCode) sessionStorage.setItem(SITE_CODE_KEY, nextSiteCode);
        }} catch (_e) {{}}
        clearRetry();
        window.location.replace((data && data.url) ? String(data.url) : nextPath);
      }} catch (e) {{
        setMsg("주차 연동 오류: " + String((e && e.message) || e));
      }}
    }})();
  </script>
</body>
</html>"""


def require_key(x_api_key: str | None):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API key")


def map_permission_to_role(permission_level: str) -> str:
    raw = (permission_level or "").strip().lower()
    if raw == "admin":
        return "admin"
    if raw == "site_admin":
        return "guard"
    if raw == "user":
        return "viewer"
    raise HTTPException(status_code=400, detail="Invalid permission_level")


def resolve_site_scope(request: Request, x_site_code: str | None = None) -> str:
    sess = read_session(request)
    if sess is not None:
        return _session_site_code(sess)
    if x_site_code:
        return normalize_site_code(x_site_code)
    raise HTTPException(status_code=400, detail="site_code is required")


def _session_site_code(sess: dict[str, Any] | None) -> str:
    raw = str((sess or {}).get("sc") or "").strip()
    if raw:
        return normalize_site_code(raw)
    raise HTTPException(status_code=403, detail="세션 단지코드가 없습니다. 시설관리에서 다시 접속하세요.")


def _is_manager_session(sess: dict[str, Any] | None) -> bool:
    role = str((sess or {}).get("r") or "").strip().lower()
    # 운영 정책: 주차 모듈은 로그인 방식과 관계없이 인증 세션이면 전체 메뉴를 사용한다.
    return role in {"admin", "guard", "viewer"}


def _require_manager_session(request: Request) -> dict[str, Any]:
    sess = read_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Login required")
    if not _is_manager_session(sess):
        raise HTTPException(status_code=403, detail="관리 권한이 필요합니다.")
    return sess


def _to_ko_status(status: str | None) -> str:
    return _STATUS_LABEL_KO.get(str(status or "").strip().lower(), "미정")


def _to_ko_verdict(verdict: str | None) -> str:
    return _VERDICT_LABEL_KO.get(str(verdict or "").strip().upper(), str(verdict or ""))


def _status_from_input(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "active"
    if raw in _STATUS_OPTIONS:
        return raw

    if raw in {"정상", "정상등록", "등록", "사용", "상시"}:
        return "active"
    if raw in {"임시", "임시등록", "temporary"}:
        return "temp"
    if raw in {"차단", "차단차량", "금지", "blocked"}:
        return "blocked"

    raise ValueError("상태 값은 active/temp/blocked 또는 정상등록/임시등록/차단차량 중 하나여야 합니다.")


def _clean_optional_text(value: Any, max_len: int) -> str | None:
    txt = str(value or "").strip()
    if not txt:
        return None
    if len(txt) > max_len:
        raise ValueError(f"텍스트 길이는 {max_len}자 이하만 가능합니다.")
    return txt


def _to_iso_date(value: Any, *, bound: str = "start") -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date fallback (1900 date system).
        serial = float(value)
        if serial > 59:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=serial)).date().isoformat()

    txt = str(value).strip()
    if not txt:
        return None

    is_end = str(bound or "").strip().lower() == "end"

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue

    for fmt in ("%Y-%m", "%Y/%m", "%Y.%m", "%Y%m"):
        try:
            dt_val = datetime.strptime(txt, fmt)
            year = int(dt_val.year)
            month = int(dt_val.month)
            day = calendar.monthrange(year, month)[1] if is_end else 1
            return date(year, month, day).isoformat()
        except ValueError:
            continue

    if re.fullmatch(r"\d{4}", txt):
        year = int(txt)
        if is_end:
            return date(year, 12, 31).isoformat()
        return date(year, 1, 1).isoformat()

    raise ValueError("날짜 형식은 YYYY-MM-DD / YYYY-MM / YYYY 중 하나로 입력하세요.")


def _vehicle_row_to_dict(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    src = dict(row)
    return {
        "site_code": src.get("site_code"),
        "plate": src.get("plate"),
        "status": src.get("status"),
        "status_ko": _to_ko_status(src.get("status")),
        "unit": src.get("unit"),
        "owner_name": src.get("owner_name"),
        "valid_from": src.get("valid_from"),
        "valid_to": src.get("valid_to"),
        "note": src.get("note"),
        "updated_at": src.get("updated_at"),
    }


def _normalize_excel_header(v: Any) -> str:
    txt = str(v or "").strip().lower()
    return re.sub(r"[^0-9a-z가-힣]", "", txt)


def _header_field_name(header: Any) -> str | None:
    normalized = _normalize_excel_header(header)
    if not normalized:
        return None
    for field, aliases in _EXCEL_HEADER_ALIASES.items():
        if normalized in {_normalize_excel_header(x) for x in aliases}:
            return field
    return None


def normalize_plate(value: str) -> str:
    return _PLATE_RE.sub("", str(value or "").upper()).strip()


def _normalize_file_extension(filename: str) -> str:
    ext = Path(str(filename or "")).suffix.lower().strip()
    if ext == ".jfif":
        return ".jpg"
    if ext == ".heif":
        return ".heic"
    return ext


def _guess_image_extension(raw: bytes) -> str | None:
    if raw.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if raw.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if len(raw) >= 12 and raw[0:4] == b"RIFF" and raw[8:12] == b"WEBP":
        return ".webp"
    if len(raw) >= 16 and raw[4:8] == b"ftyp":
        brand = raw[8:12].lower()
        if brand in {b"heic", b"heix", b"hevc", b"hevx", b"mif1", b"msf1"}:
            return ".heic"
    return None


def _validate_violation_photo_upload(photo: UploadFile) -> tuple[str, bytes]:
    ext = _normalize_file_extension(photo.filename or "")
    if ext and ext not in _ALLOWED_PHOTO_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"허용되지 않는 파일 확장자입니다: {ext}")

    content_type = str(photo.content_type or "").split(";", 1)[0].strip().lower()
    if content_type and content_type not in _ALLOWED_PHOTO_MIME_TYPES:
        raise HTTPException(status_code=400, detail=f"허용되지 않는 파일 형식입니다: {content_type}")

    raw = photo.file.read(VIOLATION_UPLOAD_MAX_BYTES + 1)
    if not raw:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
    if len(raw) > VIOLATION_UPLOAD_MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"업로드 크기 제한({VIOLATION_UPLOAD_MAX_BYTES} bytes)을 초과했습니다.")

    guessed_ext = _guess_image_extension(raw)
    if not guessed_ext:
        raise HTTPException(status_code=400, detail="지원하지 않는 이미지 포맷입니다.")

    if ext:
        normalized_ext = ".jpg" if ext in {".jpg", ".jpeg"} else ext
        if normalized_ext != guessed_ext:
            raise HTTPException(status_code=400, detail="파일 확장자와 실제 이미지 형식이 일치하지 않습니다.")
    else:
        ext = guessed_ext

    if ext == ".jpeg":
        ext = ".jpg"
    if ext == ".heif":
        ext = ".heic"
    return ext, raw


def _raise_db_error(exc: sqlite3.Error, *, action: str) -> None:
    msg = str(exc or "").strip()
    lowered = msg.lower()
    if "database is locked" in lowered or "database is busy" in lowered:
        raise HTTPException(
            status_code=503,
            detail=f"{action} 중 DB가 사용 중입니다. 잠시 후 다시 시도하세요.",
        ) from exc
    if isinstance(exc, sqlite3.IntegrityError):
        raise HTTPException(
            status_code=409,
            detail=f"{action} 중 중복 또는 무결성 오류가 발생했습니다. ({msg})",
        ) from exc
    raise HTTPException(
        status_code=500,
        detail=f"{action} 중 DB 오류가 발생했습니다. ({msg})",
    ) from exc


def check_plate_record(site_code: str, plate: str) -> "CheckResponse":
    p = normalize_plate(plate)
    if len(p) < 4:
        raise HTTPException(status_code=400, detail="plate is too short")

    with connect() as con:
        row = con.execute("SELECT * FROM vehicles WHERE site_code = ? AND plate = ?", (site_code, p)).fetchone()

    if not row:
        return CheckResponse(site_code=site_code, plate=p, verdict="UNREGISTERED", message="미등록 차량")

    status = (row["status"] or "active").lower()
    vf, vt = row["valid_from"], row["valid_to"]
    today = today_iso()

    if status == "blocked":
        return CheckResponse(
            site_code=site_code,
            plate=p,
            verdict="BLOCKED",
            message="차단 차량",
            unit=row["unit"],
            owner_name=row["owner_name"],
            status=status,
            valid_from=vf,
            valid_to=vt,
        )
    if vt and today > vt:
        return CheckResponse(
            site_code=site_code,
            plate=p,
            verdict="EXPIRED",
            message="기간 만료",
            unit=row["unit"],
            owner_name=row["owner_name"],
            status=status,
            valid_from=vf,
            valid_to=vt,
        )
    if status == "temp":
        return CheckResponse(
            site_code=site_code,
            plate=p,
            verdict="TEMP",
            message="임시 등록",
            unit=row["unit"],
            owner_name=row["owner_name"],
            status=status,
            valid_from=vf,
            valid_to=vt,
        )
    return CheckResponse(
        site_code=site_code,
        plate=p,
        verdict="OK",
        message="정상 등록",
        unit=row["unit"],
        owner_name=row["owner_name"],
        status=status,
        valid_from=vf,
        valid_to=vt,
    )

@app.on_event("startup")
def _startup():
    init_db()
    seed_demo()
    seed_users()

def today_iso() -> str:
    return date.today().isoformat()

@app.get("/health")
def health():
    return {"ok": True}

@app.get("/")
def home():
    return RedirectResponse(url=app_url("/admin2"), status_code=302)

@app.get("/login", response_class=HTMLResponse)
def login_page():
    if not LOCAL_LOGIN_ENABLED:
        return integration_required_page(status_code=200)
    login_action = app_url("/login")
    return f"<h2>Login</h2><form method='POST' action='{login_action}'><input name='username'/><input name='password' type='password'/><button>Login</button></form>"

@app.post("/login")
def login_submit(username: str = Form(...), password: str = Form(...)):
    if not LOCAL_LOGIN_ENABLED:
        return integration_required_page(status_code=403)
    u = username.strip()
    with connect() as con:
        row = con.execute("SELECT * FROM users WHERE username=?", (u,)).fetchone()
    if not row or not pbkdf2_verify(password, row["pw_hash"]):
        return HTMLResponse("Invalid credentials", status_code=401)
    token = make_session(u, row["role"], site_code=DEFAULT_SITE_CODE)
    resp = RedirectResponse(url=app_url("/admin2"), status_code=302)
    resp.set_cookie("parking_session", token, httponly=True, secure=COOKIE_SECURE, samesite="lax", path=ROOT_PATH or "/")
    return resp


@app.get("/sso")
def sso_login(ctx: str):
    try:
        payload = _ctx_ser.loads(ctx, max_age=CONTEXT_MAX_AGE)
    except SignatureExpired as exc:
        raise HTTPException(status_code=401, detail="Context token expired") from exc
    except BadSignature as exc:
        raise HTTPException(status_code=401, detail="Invalid context token") from exc

    raw_site_code = str(payload.get("site_code") or "").strip()
    if not raw_site_code:
        raise HTTPException(status_code=400, detail="Context token missing site_code")
    site_code = normalize_site_code(raw_site_code)

    permission_level = str(payload.get("permission_level") or "").strip().lower()
    if not permission_level:
        raise HTTPException(status_code=400, detail="Context token missing permission_level")
    role = map_permission_to_role(permission_level)
    session_user = str(payload.get("login_id") or payload.get("user_login") or "ka-part-user").strip() or "ka-part-user"
    extras = {
        "display_name": str(payload.get("user_name") or "").strip(),
        "site_name": str(payload.get("site_name") or "").strip(),
    }
    portal_level = str(payload.get("portal_permission_level") or "").strip().lower()
    if portal_level:
        extras["portal_permission_level"] = portal_level
    token = make_session(session_user, role, site_code=site_code, extras=extras)
    resp = RedirectResponse(url=app_url("/admin2"), status_code=302)
    resp.set_cookie("parking_session", token, httponly=True, secure=COOKIE_SECURE, samesite="lax", path=ROOT_PATH or "/")
    return resp

@app.post("/logout")
def logout():
    target = app_url("/login")
    if (not LOCAL_LOGIN_ENABLED) and PORTAL_URL:
        target = PORTAL_URL
    resp = RedirectResponse(url=target, status_code=302)
    resp.delete_cookie("parking_session", path=ROOT_PATH or "/", secure=COOKIE_SECURE, httponly=True, samesite="lax")
    return resp

class CheckResponse(BaseModel):
    site_code: str
    plate: str
    verdict: Literal["OK","UNREGISTERED","BLOCKED","EXPIRED","TEMP"]
    message: str
    unit: Optional[str] = None
    owner_name: Optional[str] = None
    status: Optional[str] = None
    valid_from: Optional[str] = None
    valid_to: Optional[str] = None

@app.get("/api/plates/check", response_model=CheckResponse)
def check_plate(
    plate: str,
    request: Request,
    x_api_key: str | None = Header(default=None, alias="X-API-Key"),
    x_site_code: str | None = Header(default=None, alias="X-Site-Code"),
):
    require_key(x_api_key)
    site_code = resolve_site_scope(request, x_site_code)
    return check_plate_record(site_code, plate)


@app.get("/api/session/plates/check", response_model=CheckResponse)
def check_plate_session(
    plate: str,
    request: Request,
):
    sess = read_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Login required")
    site_code = _session_site_code(sess)
    return check_plate_record(site_code, plate)

class ViolationOut(BaseModel):
    id: int
    site_code: str
    plate: str
    verdict: str
    rule_code: Optional[str] = None
    location: Optional[str] = None
    memo: Optional[str] = None
    inspector: Optional[str] = None
    photo_path: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None
    created_at: str


class SessionViolationIn(BaseModel):
    plate: str
    verdict: Literal["OK", "UNREGISTERED", "BLOCKED", "EXPIRED", "TEMP"] = "UNREGISTERED"
    location: Optional[str] = None
    memo: Optional[str] = None


class VehicleUpsertIn(BaseModel):
    plate: str
    status: str = "active"
    unit: Optional[str] = None
    owner_name: Optional[str] = None
    valid_from: Optional[str] = None
    valid_to: Optional[str] = None
    note: Optional[str] = None


def _normalize_vehicle_payload(payload: VehicleUpsertIn) -> dict[str, Any]:
    plate = normalize_plate(payload.plate)
    if len(plate) < 4:
        raise HTTPException(status_code=400, detail="차량번호 형식이 올바르지 않습니다.")

    try:
        status = _status_from_input(payload.status)
        unit = _clean_optional_text(payload.unit, 40)
        owner_name = _clean_optional_text(payload.owner_name, 60)
        valid_from = _to_iso_date(payload.valid_from, bound="start")
        valid_to = _to_iso_date(payload.valid_to, bound="end")
        note = _clean_optional_text(payload.note, 200)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if valid_from and valid_to and valid_from > valid_to:
        raise HTTPException(status_code=400, detail="적용시작일은 적용종료일보다 클 수 없습니다.")

    return {
        "plate": plate,
        "status": status,
        "unit": unit,
        "owner_name": owner_name,
        "valid_from": valid_from,
        "valid_to": valid_to,
        "note": note,
    }


@app.post("/api/violations/upload", response_model=ViolationOut)
def create_violation_with_photo(
    request: Request,
    plate: str = Form(...),
    verdict: str = Form(...),
    rule_code: str | None = Form(None),
    location: str | None = Form(None),
    memo: str | None = Form(None),
    inspector: str | None = Form(None),
    lat: float | None = Form(None),
    lng: float | None = Form(None),
    photo: UploadFile = File(...),
    x_api_key: str | None = Header(default=None, alias="X-API-Key"),
    x_site_code: str | None = Header(default=None, alias="X-Site-Code"),
):
    require_key(x_api_key)
    site_code = resolve_site_scope(request, x_site_code)
    p = normalize_plate(plate)
    if len(p) < 4:
        raise HTTPException(status_code=400, detail="차량번호 형식이 올바르지 않습니다.")
    verdict_value = str(verdict or "").strip().upper()
    if verdict_value not in _VERDICT_OPTIONS:
        raise HTTPException(status_code=400, detail="verdict 값이 올바르지 않습니다.")
    ext, raw = _validate_violation_photo_upload(photo)
    fname = f"{uuid.uuid4().hex}{ext}"
    fpath = UPLOAD_DIR / fname
    with open(fpath, "wb") as f:
        f.write(raw)
    rel = app_url(f"/uploads/{fname}")
    try:
        with connect() as con:
            cur = con.execute(
                "INSERT INTO violations (site_code, plate, verdict, rule_code, location, memo, inspector, photo_path, lat, lng) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (site_code, p, verdict_value, rule_code, location, memo, inspector, rel, lat, lng),
            )
            vid = cur.lastrowid
            row = con.execute("SELECT * FROM violations WHERE id = ?", (vid,)).fetchone()
    except sqlite3.Error as exc:
        try:
            fpath.unlink(missing_ok=True)
        except Exception:
            pass
        _raise_db_error(exc, action="위반기록 저장")
    return ViolationOut(**dict(row))


@app.post("/api/session/violations", response_model=ViolationOut)
def create_violation_session(request: Request, payload: SessionViolationIn):
    sess = read_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Login required")

    site_code = _session_site_code(sess)
    plate = normalize_plate(payload.plate)
    if len(plate) < 4:
        raise HTTPException(status_code=400, detail="plate is too short")

    verdict = str(payload.verdict or "UNREGISTERED").strip().upper()
    if verdict not in _VERDICT_OPTIONS:
        raise HTTPException(status_code=400, detail="invalid verdict")

    inspector = str(sess.get("display_name") or sess.get("u") or "ka-part-user").strip() or "ka-part-user"
    try:
        with connect() as con:
            cur = con.execute(
                "INSERT INTO violations (site_code, plate, verdict, location, memo, inspector) VALUES (?,?,?,?,?,?)",
                (site_code, plate, verdict, payload.location, payload.memo, inspector),
            )
            vid = cur.lastrowid
            row = con.execute("SELECT * FROM violations WHERE id = ?", (vid,)).fetchone()
    except sqlite3.Error as exc:
        _raise_db_error(exc, action="위반기록 저장")
    return ViolationOut(**dict(row))


@app.get("/api/session/violations/recent")
def list_violations_session(request: Request, limit: int = 20):
    sess = read_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Login required")

    site_code = _session_site_code(sess)
    use_limit = max(1, min(int(limit), 100))
    with connect() as con:
        rows = con.execute(
            """
            SELECT id, site_code, plate, verdict, rule_code, location, memo, inspector,
                   photo_path, lat, lng, created_at
            FROM violations
            WHERE site_code=?
            ORDER BY id DESC
            LIMIT ?
            """,
            (site_code, use_limit),
        ).fetchall()
    items = []
    for r in rows:
        d = dict(r)
        d["verdict_ko"] = _to_ko_verdict(d.get("verdict"))
        items.append(d)
    return {"ok": True, "items": items}


@app.get("/api/session/vehicles")
def list_vehicles_session(request: Request, q: str = "", limit: int = 200):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    use_limit = max(1, min(int(limit), 1000))
    q_plate = normalize_plate(q)
    q_text = str(q or "").strip()

    with connect() as con:
        if q_text:
            rows = con.execute(
                """
                SELECT site_code, plate, status, unit, owner_name, valid_from, valid_to, note, updated_at
                FROM vehicles
                WHERE site_code = ?
                  AND (
                    plate LIKE ?
                    OR IFNULL(owner_name, '') LIKE ?
                    OR IFNULL(unit, '') LIKE ?
                  )
                ORDER BY updated_at DESC, plate ASC
                LIMIT ?
                """,
                (
                    site_code,
                    f"%{q_plate or q_text.upper()}%",
                    f"%{q_text}%",
                    f"%{q_text}%",
                    use_limit,
                ),
            ).fetchall()
        else:
            rows = con.execute(
                """
                SELECT site_code, plate, status, unit, owner_name, valid_from, valid_to, note, updated_at
                FROM vehicles
                WHERE site_code = ?
                ORDER BY updated_at DESC, plate ASC
                LIMIT ?
                """,
                (site_code, use_limit),
            ).fetchall()

    return {"ok": True, "items": [_vehicle_row_to_dict(r) for r in rows], "count": len(rows)}


@app.get("/api/session/system_check")
def system_check_session(request: Request):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    role = str(sess.get("r") or "").strip().lower() or "unknown"

    try:
        with connect() as con:
            jm_row = con.execute("PRAGMA journal_mode").fetchone()
            bt_row = con.execute("PRAGMA busy_timeout").fetchone()
            cols = con.execute("PRAGMA table_info(vehicles)").fetchall()
            pk_cols = sorted(
                [
                    (int(c["pk"] or 0), str(c["name"]))
                    for c in cols
                    if int(c["pk"] or 0) > 0
                ],
                key=lambda x: x[0],
            )
            cnt_row = con.execute(
                "SELECT COUNT(*) AS n FROM vehicles WHERE site_code=?",
                (site_code,),
            ).fetchone()

            # write test without persistence
            probe_plate = f"CHECK{uuid.uuid4().hex[:8].upper()}"
            con.execute("SAVEPOINT sp_system_check")
            con.execute(
                """
                INSERT INTO vehicles(site_code, plate, status, updated_at)
                VALUES(?,?,?, datetime('now'))
                """,
                (site_code, probe_plate, "temp"),
            )
            con.execute(
                "DELETE FROM vehicles WHERE site_code=? AND plate=?",
                (site_code, probe_plate),
            )
            con.execute("RELEASE sp_system_check")

        return {
            "ok": True,
            "role": role,
            "site_code": site_code,
            "db_path": str(DB_PATH),
            "journal_mode": (jm_row[0] if jm_row else None),
            "busy_timeout_ms": int(bt_row[0]) if bt_row and bt_row[0] is not None else None,
            "vehicles_count_for_site": int(cnt_row["n"]) if cnt_row else 0,
            "vehicles_pk": [name for _pk, name in pk_cols],
            "can_write": True,
        }
    except sqlite3.Error as exc:
        msg = str(exc or "").strip()
        return {
            "ok": False,
            "role": role,
            "site_code": site_code,
            "db_path": str(DB_PATH),
            "can_write": False,
            "detail": msg or "sqlite error",
        }


@app.post("/api/session/vehicles")
def create_vehicle_session(request: Request, payload: VehicleUpsertIn):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    data = _normalize_vehicle_payload(payload)

    try:
        with connect() as con:
            exists = con.execute(
                "SELECT 1 FROM vehicles WHERE site_code=? AND plate=?",
                (site_code, data["plate"]),
            ).fetchone()
            if exists:
                raise HTTPException(status_code=409, detail="이미 등록된 차량번호입니다. 수정 기능을 사용하세요.")

            con.execute(
                """
                INSERT INTO vehicles(site_code, plate, unit, owner_name, status, valid_from, valid_to, note, updated_at)
                VALUES(?,?,?,?,?,?,?, ?, datetime('now'))
                """,
                (
                    site_code,
                    data["plate"],
                    data["unit"],
                    data["owner_name"],
                    data["status"],
                    data["valid_from"],
                    data["valid_to"],
                    data["note"],
                ),
            )
            row = con.execute(
                """
                SELECT site_code, plate, status, unit, owner_name, valid_from, valid_to, note, updated_at
                FROM vehicles
                WHERE site_code=? AND plate=?
                """,
                (site_code, data["plate"]),
            ).fetchone()
    except sqlite3.Error as exc:
        _raise_db_error(exc, action="신규등록")
    return {"ok": True, "item": _vehicle_row_to_dict(row)}


@app.put("/api/session/vehicles/{plate}")
def update_vehicle_session(plate: str, request: Request, payload: VehicleUpsertIn):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    target_plate = normalize_plate(plate)
    if len(target_plate) < 4:
        raise HTTPException(status_code=400, detail="차량번호 형식이 올바르지 않습니다.")

    data = _normalize_vehicle_payload(payload)
    if data["plate"] != target_plate:
        raise HTTPException(status_code=400, detail="수정 시 차량번호는 기존 차량번호와 같아야 합니다.")

    try:
        with connect() as con:
            exists = con.execute(
                "SELECT 1 FROM vehicles WHERE site_code=? AND plate=?",
                (site_code, target_plate),
            ).fetchone()
            if not exists:
                raise HTTPException(status_code=404, detail="수정할 차량을 찾지 못했습니다.")

            con.execute(
                """
                UPDATE vehicles
                SET unit=?, owner_name=?, status=?, valid_from=?, valid_to=?, note=?, updated_at=datetime('now')
                WHERE site_code=? AND plate=?
                """,
                (
                    data["unit"],
                    data["owner_name"],
                    data["status"],
                    data["valid_from"],
                    data["valid_to"],
                    data["note"],
                    site_code,
                    target_plate,
                ),
            )
            row = con.execute(
                """
                SELECT site_code, plate, status, unit, owner_name, valid_from, valid_to, note, updated_at
                FROM vehicles
                WHERE site_code=? AND plate=?
                """,
                (site_code, target_plate),
            ).fetchone()
    except sqlite3.Error as exc:
        _raise_db_error(exc, action="차량정보 수정")
    return {"ok": True, "item": _vehicle_row_to_dict(row)}


@app.delete("/api/session/vehicles/{plate}")
def delete_vehicle_session(plate: str, request: Request):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    target_plate = normalize_plate(plate)
    if len(target_plate) < 4:
        raise HTTPException(status_code=400, detail="차량번호 형식이 올바르지 않습니다.")

    try:
        with connect() as con:
            exists = con.execute(
                "SELECT 1 FROM vehicles WHERE site_code=? AND plate=?",
                (site_code, target_plate),
            ).fetchone()
            if not exists:
                raise HTTPException(status_code=404, detail="삭제할 차량을 찾지 못했습니다.")

            con.execute(
                "DELETE FROM vehicles WHERE site_code=? AND plate=?",
                (site_code, target_plate),
            )
    except sqlite3.Error as exc:
        _raise_db_error(exc, action="차량정보 삭제")
    return {"ok": True, "plate": target_plate}


@app.post("/api/session/vehicles/import_excel")
async def import_vehicles_excel_session(request: Request, file: UploadFile = File(...)):
    sess = _require_manager_session(request)
    site_code = _session_site_code(sess)
    filename = str(file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드할 수 있습니다.")

    raw = await file.read(EXCEL_UPLOAD_MAX_BYTES + 1)
    if not raw:
        raise HTTPException(status_code=400, detail="업로드된 파일이 비어 있습니다.")
    if len(raw) > EXCEL_UPLOAD_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"엑셀 업로드 크기 제한({EXCEL_UPLOAD_MAX_BYTES} bytes)을 초과했습니다.",
        )

    try:
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"엑셀 파일을 읽을 수 없습니다: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="엑셀 시트에 데이터가 없습니다.")

    headers = list(rows[0] or [])
    idx_field: dict[int, str] = {}
    unknown_headers: list[str] = []
    for idx, h in enumerate(headers):
        field = _header_field_name(h)
        if field and field not in idx_field.values():
            idx_field[idx] = field
        elif str(h or "").strip():
            unknown_headers.append(str(h).strip())

    if "plate" not in idx_field.values():
        return JSONResponse(
            status_code=200,
            content={
                "ok": False,
                "message": "필수 컬럼 '차량번호'를 찾지 못했습니다.",
                "guidance": {
                    "required_columns": ["차량번호(필수)"],
                    "optional_columns": ["상태", "동호수", "소유자", "적용시작일", "적용종료일", "비고"],
                    "status_guide": "상태는 정상등록/임시등록/차단차량 또는 active/temp/blocked 값을 사용하세요.",
                    "date_guide": "날짜는 YYYY-MM-DD / YYYY-MM / YYYY 형식을 지원합니다. (월/년 입력 시 시작일=1일, 종료일=말일 자동보정)",
                },
                "unknown_headers": unknown_headers,
            },
        )

    parsed: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    skipped_empty = 0

    for row_index, row in enumerate(rows[1:], start=2):
        if not row or all((v is None) or (str(v).strip() == "") for v in row):
            skipped_empty += 1
            continue

        raw_item: dict[str, Any] = {}
        for idx, field in idx_field.items():
            raw_item[field] = row[idx] if idx < len(row) else None

        try:
            plate = normalize_plate(raw_item.get("plate"))
            if len(plate) < 4:
                raise ValueError("차량번호 형식이 올바르지 않습니다.")
            status = _status_from_input(raw_item.get("status"))
            unit = _clean_optional_text(raw_item.get("unit"), 40)
            owner_name = _clean_optional_text(raw_item.get("owner_name"), 60)
            valid_from = _to_iso_date(raw_item.get("valid_from"), bound="start")
            valid_to = _to_iso_date(raw_item.get("valid_to"), bound="end")
            note = _clean_optional_text(raw_item.get("note"), 200)

            if valid_from and valid_to and valid_from > valid_to:
                raise ValueError("적용시작일은 적용종료일보다 클 수 없습니다.")

            parsed.append(
                {
                    "plate": plate,
                    "status": status,
                    "unit": unit,
                    "owner_name": owner_name,
                    "valid_from": valid_from,
                    "valid_to": valid_to,
                    "note": note,
                }
            )
        except ValueError as exc:
            errors.append(
                {
                    "row": row_index,
                    "message": str(exc),
                    "sample": {
                        "차량번호": str(raw_item.get("plate") or ""),
                        "상태": str(raw_item.get("status") or ""),
                        "동호수": str(raw_item.get("unit") or ""),
                    },
                }
            )

    inserted = 0
    updated = 0
    try:
        with connect() as con:
            for item in parsed:
                exists = con.execute(
                    "SELECT 1 FROM vehicles WHERE site_code=? AND plate=?",
                    (site_code, item["plate"]),
                ).fetchone()
                con.execute(
                    """
                    INSERT INTO vehicles(site_code, plate, unit, owner_name, status, valid_from, valid_to, note, updated_at)
                    VALUES(?,?,?,?,?,?,?, ?, datetime('now'))
                    ON CONFLICT(site_code, plate) DO UPDATE SET
                      unit=excluded.unit,
                      owner_name=excluded.owner_name,
                      status=excluded.status,
                      valid_from=excluded.valid_from,
                      valid_to=excluded.valid_to,
                      note=excluded.note,
                      updated_at=datetime('now')
                    """,
                    (
                        site_code,
                        item["plate"],
                        item["unit"],
                        item["owner_name"],
                        item["status"],
                        item["valid_from"],
                        item["valid_to"],
                        item["note"],
                    ),
                )
                if exists:
                    updated += 1
                else:
                    inserted += 1
    except sqlite3.Error as exc:
        _raise_db_error(exc, action="엑셀 가져오기")

    return {
        "ok": len(errors) == 0,
        "message": "엑셀 가져오기를 완료했습니다." if not errors else "일부 행에서 형식 오류가 있어 안내를 확인하세요.",
        "inserted": inserted,
        "updated": updated,
        "skipped_empty": skipped_empty,
        "error_count": len(errors),
        "errors": errors[:30],
        "guidance": {
            "required_columns": ["차량번호(필수)"],
            "optional_columns": ["상태", "동호수", "소유자", "적용시작일", "적용종료일", "비고"],
            "status_guide": "상태는 정상등록/임시등록/차단차량 또는 active/temp/blocked 값을 사용하세요.",
            "date_guide": "날짜는 YYYY-MM-DD / YYYY-MM / YYYY 형식을 지원합니다. (월/년 입력 시 시작일=1일, 종료일=말일 자동보정)",
        },
        "unknown_headers": unknown_headers,
    }


def esc(v): return _html.escape(str(v)) if v is not None else ""


ADMIN2_HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover" />
  <title>주차관리 DB</title>
  <style>
    :root { --bg:#0a1a23; --panel:#112735; --line:#27465a; --text:#e9f2f7; --muted:#8fb0c2; --ok:#16a34a; --warn:#b45309; --bad:#dc2626; --btn:#1d4ed8; }
    * { box-sizing:border-box; }
    html, body { max-width:100%; overflow-x:hidden; }
    body { margin:0; font-family:"Noto Sans KR",system-ui,sans-serif; color:var(--text); background:radial-gradient(circle at top, #173a4e, #0a1a23 58%); }
    .wrap { width:100%; max-width:1100px; margin:0 auto; padding:16px; padding: calc(16px + env(safe-area-inset-top)) calc(16px + env(safe-area-inset-right)) calc(16px + env(safe-area-inset-bottom)) calc(16px + env(safe-area-inset-left)); display:grid; gap:12px; }
    .card { background:rgba(17,39,53,.94); border:1px solid var(--line); border-radius:14px; padding:14px; }
    .head { display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap; }
    .title { margin:0; font-size:22px; }
    .sub { color:var(--muted); font-size:13px; }
    h2 { margin:0; font-size:19px; }
    .nav { display:flex; gap:8px; flex-wrap:wrap; }
    .btn { border:1px solid var(--line); background:var(--btn); color:#fff; border-radius:10px; padding:9px 12px; font-weight:700; cursor:pointer; text-decoration:none; }
    .btn.ghost { background:transparent; color:var(--text); }
    .btn.warn { background:var(--warn); border-color:#92400e; }
    .btn.bad { background:var(--bad); border-color:#7f1d1d; }
    .grid { display:grid; grid-template-columns:1fr 1fr; gap:12px; }
    .row { display:flex; gap:8px; flex-wrap:wrap; margin-top:8px; }
    .field { display:grid; gap:6px; width:100%; }
    .field.w50 { width:calc(50% - 4px); }
    label { color:var(--muted); font-size:12px; }
    input, select { width:100%; border:1px solid var(--line); border-radius:10px; background:#0d2230; color:var(--text); padding:9px 11px; }
    table { width:100%; border-collapse:collapse; margin-top:8px; font-size:13px; }
    th, td { border:1px solid var(--line); padding:7px 8px; text-align:left; vertical-align:top; }
    th { background:#0f2431; color:#cae2ef; }
    .table-scroll { border:1px solid var(--line); border-radius:10px; overflow-x:auto; overflow-y:visible; margin-top:8px; max-width:100%; }
    .table-scroll table { margin-top:0; border-collapse:separate; border-spacing:0; min-width:920px; }
    .table-scroll th, .table-scroll td { border-right:0; border-bottom:0; }
    .table-scroll tr > *:last-child { border-right:1px solid var(--line); }
    .table-scroll tbody tr:last-child > * { border-bottom:1px solid var(--line); }
    #tbVehicles thead th { position: sticky; top: 0; z-index: 2; }
    #tbViolations thead th { position: sticky; top: 0; z-index: 2; }
    #violationsTableWrap table { min-width:760px; }
    tr.sel { background:rgba(29,78,216,.25); }
    .msg { min-height:22px; margin-top:8px; color:#b8d7e8; white-space:pre-wrap; }
    .msg.ok { color:#8ee6a6; }
    .msg.err { color:#ffb4b4; }
    .hintbox { border:1px dashed var(--line); background:#0c2130; border-radius:10px; padding:10px; font-size:12px; color:var(--muted); line-height:1.6; }
    .perm { color:#ffd4a8; font-size:13px; }
    .tab-section { }

    details.adv { margin-top:8px; border:1px dashed var(--line); background:#0c2130; border-radius:12px; padding:10px; }
    details.adv > summary { cursor:pointer; font-weight:800; color:#cae2ef; }
    details.adv > summary::-webkit-details-marker { display:none; }
    details.adv > summary:before { content:"▸"; display:inline-block; margin-right:8px; color:var(--muted); }
    details.adv[open] > summary:before { content:"▾"; }
    details.adv[open] > summary { margin-bottom:8px; }

    .bottom-tabs { display:none; position:fixed; left:0; right:0; bottom:0; z-index:999; background:rgba(10,26,35,.82); border-top:1px solid var(--line); -webkit-backdrop-filter: blur(10px); backdrop-filter: blur(10px); padding:8px 8px calc(8px + env(safe-area-inset-bottom)); gap:8px; }
    .bottom-tabs .tab { flex:1; border:1px solid var(--line); border-radius:12px; background:rgba(13,34,48,.85); color:var(--text); padding:10px 8px; font-weight:900; font-size:12px; text-decoration:none; cursor:pointer; text-align:center; }
    .bottom-tabs .tab.active { border-color:rgba(29,78,216,.75); background:rgba(29,78,216,.25); }
    .bottom-tabs .tab.danger { color:#ffd4d4; border-color:#7f1d1d; background:rgba(220,38,38,.12); }
    .bottom-tabs .tab.disabled { opacity:.55; pointer-events:none; }
    @media (max-width: 900px) {
      body.tabbed .wrap { padding-bottom: calc(92px + env(safe-area-inset-bottom)); }
      body.tabbed .bottom-tabs { display:flex; }
      body.tabbed .tab-section { display:none; }
      body.tabbed .tab-section.active { display:block; }
      .grid { grid-template-columns:1fr; }
      .field.w50 { width:100%; }
      input, select { font-size:16px; }
      .nav { width:100%; display:grid; grid-template-columns:1fr 1fr; }
      .nav .btn { width:100%; text-align:center; }
      .title { font-size:18px; }
      h2 { font-size:16px; }
      .row.actions { display:grid; grid-template-columns:1fr 1fr; }
      .row.actions .btn { width:100%; }
      .table-scroll { -webkit-overflow-scrolling: touch; }
      .table-scroll table { min-width:0; }
      #violationsTableWrap table { min-width:0; }
      table { font-size:12px; }

      /* 모바일에서는 중요한 컬럼만 남겨 가로 잘림/가로스크롤을 최소화 */
      /* 차량: 차량번호/상태/동호수/적용종료일 중심 */
      #tbVehicles th:nth-child(4), #tbVehicles td:nth-child(4),
      #tbVehicles th:nth-child(5), #tbVehicles td:nth-child(5),
      #tbVehicles th:nth-child(7), #tbVehicles td:nth-child(7),
      #tbVehicles th:nth-child(8), #tbVehicles td:nth-child(8) { display:none; }

      /* 위반: 발생시각/차량번호/판정/사진 중심 */
      #tbViolations th:nth-child(4), #tbViolations td:nth-child(4),
      #tbViolations th:nth-child(5), #tbViolations td:nth-child(5),
      #tbViolations th:nth-child(6), #tbViolations td:nth-child(6) { display:none; }
    }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="card">
      <div class="head">
        <div>
          <h1 class="title">주차관리 DB</h1>
          <div class="sub" id="ctxLine"></div>
        </div>
        <div class="nav">
          <a id="btnBack" class="btn ghost" href="/pwa/">시설관리로 돌아가기</a>
          <a id="btnScanner" class="btn ghost" href="./scanner">번호판 스캐너</a>
          <button id="btnRefreshAll" class="btn ghost" type="button">새로고침</button>
          <button id="btnLogout" class="btn ghost" type="button">로그아웃</button>
        </div>
      </div>
      <div class="perm" id="permLine"></div>
      <div class="sub" id="sysLine"></div>
    </section>

    <section id="sectionVehicles" class="card tab-section">
      <h2 style="margin:0;">차량정보 DB 관리</h2>
      <div class="hintbox" id="importGuide">
        엑셀 컬럼 안내
        - 필수: 차량번호
        - 선택: 상태, 동호수, 소유자, 적용시작일, 적용종료일, 비고
        - 상태값: 정상등록/임시등록/차단차량 (또는 active/temp/blocked)
        - 날짜: YYYY-MM-DD / YYYY-MM / YYYY 지원
          (월/년 입력 시 시작일=1일, 종료일=해당 월/연도 말일 자동보정)
      </div>
      <div id="managerOnly">
        <div class="row">
          <input type="file" id="excelFile" accept=".xlsx" />
          <button id="btnImportExcel" class="btn warn" type="button">엑셀 불러오기</button>
        </div>
        <div class="grid">
          <div>
            <div class="row">
              <div class="field w50">
                <label>차량번호</label>
                <input id="fPlate" placeholder="예: 123가4567" />
              </div>
              <div class="field w50">
                <label>등록상태</label>
                <select id="fStatus">
                  <option value="active">정상등록</option>
                  <option value="temp">임시등록</option>
                  <option value="blocked">차단차량</option>
                </select>
              </div>
            </div>

            <details id="advFields" class="adv" open>
              <summary>추가 입력 (선택)</summary>
              <div class="row">
                <div class="field w50">
                  <label>동호수</label>
                  <input id="fUnit" placeholder="예: 101-1203" />
                </div>
                <div class="field w50">
                  <label>소유자</label>
                  <input id="fOwner" placeholder="예: 홍길동" />
                </div>
                <div class="field w50">
                  <label>적용시작일 (일/월/년)</label>
                  <input id="fFrom" placeholder="예: 2026-02-11 / 2026-02 / 2026" />
                </div>
                <div class="field w50">
                  <label>적용종료일 (일/월/년)</label>
                  <input id="fTo" placeholder="예: 2026-12-31 / 2026-12 / 2026" />
                </div>
                <div class="field">
                  <label>비고</label>
                  <input id="fNote" placeholder="선택 입력" />
                </div>
              </div>
            </details>

            <div class="row actions">
              <button id="btnCreate" class="btn" type="button">신규등록</button>
              <button id="btnUpdate" class="btn warn" type="button">선택수정</button>
              <button id="btnDelete" class="btn bad" type="button">선택삭제</button>
              <button id="btnClear" class="btn ghost" type="button">입력초기화</button>
            </div>
          </div>
          <div>
            <div class="field">
              <label>검색 (차량번호/동호수/소유자)</label>
              <input id="qVehicle" placeholder="검색어 입력 후 Enter" />
            </div>
            <div class="msg" id="msgVehicle"></div>
          </div>
        </div>
      </div>
      <div class="msg" id="msgImport"></div>
      <div id="vehicleTableWrap" class="table-scroll">
        <table id="tbVehicles">
          <thead>
            <tr>
              <th>차량번호</th>
              <th>등록상태</th>
              <th>동호수</th>
              <th>소유자</th>
              <th>적용시작일</th>
              <th>적용종료일</th>
              <th>비고</th>
              <th>갱신시각</th>
            </tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </section>

    <section id="sectionViolations" class="card tab-section">
      <h2 style="margin:0;">위반기록</h2>
      <div class="row">
        <button id="btnReloadViolations" class="btn ghost" type="button">위반기록 새로고침</button>
      </div>
      <div id="violationsTableWrap" class="table-scroll">
        <table id="tbViolations">
          <thead>
            <tr>
              <th>발생시각</th>
              <th>차량번호</th>
              <th>판정</th>
              <th>위치</th>
              <th>메모</th>
              <th>등록자</th>
              <th>사진</th>
            </tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </section>
  </main>

  <nav class="bottom-tabs" id="bottomTabs" aria-label="빠른 메뉴">
    <button type="button" class="tab" id="tabVehicles">차량DB</button>
    <button type="button" class="tab" id="tabViolations">위반기록</button>
    <a class="tab" id="tabScanner" href="./scanner">스캐너</a>
    <button type="button" class="tab danger" id="tabLogout">로그아웃</button>
  </nav>

  <script>window.__ADMIN_BOOT__ = __ADMIN_BOOT_JSON__;</script>
  <script>
  (function(){
    const boot = window.__ADMIN_BOOT__ || {};
    const $ = (id) => document.getElementById(id);
    const qs = (selector) => document.querySelector(selector);
    const _inferredApiBase = (function(){
      if (boot.api_base) return String(boot.api_base);
      const scanner = String(boot.scanner_url || "");
      const m = scanner.match(/^(.*)\/scanner\/?$/);
      if (m && m[1]) return `${m[1]}/api/session`;
      return "/parking/api/session";
    })();
    const API_BASE = String(_inferredApiBase || "/parking/api/session");
    function apiUrl(path){
      const cleaned = String(path || "").replace(/^\/+/, "");
      if (!cleaned) return API_BASE;
      if (/^https?:\/\//i.test(cleaned)) return cleaned;
      if (cleaned.startsWith("/")) return cleaned;
      if (API_BASE.endsWith("/")) return `${API_BASE}${cleaned}`;
      return `${API_BASE}/${cleaned}`;
    }
    function on(id, eventName, handler){
      const el = $(id);
      if (!el) return false;
      el.addEventListener(eventName, handler);
      return true;
    }
    const state = { selectedPlate: "", vehicles: [] };
    const isManager = !!boot.is_manager;

    function setMsg(id, text, ok) {
      const el = $(id);
      if (!el) return;
      el.textContent = text || "";
      el.className = "msg" + (text ? (ok ? " ok" : " err") : "");
    }

    function friendlyErrMsg(error){
      const msg = String((error && error.message) || error || "").trim();
      if (!msg) return "알 수 없는 오류";
      if (msg.includes("이미 등록된 차량번호")) {
        return "이미 등록된 차량번호입니다. 기존 행을 선택해 '선택수정'을 사용하세요.";
      }
      if (msg.includes("DB가 사용 중")) {
        return "DB 사용 중으로 처리 지연이 발생했습니다. 3~5초 후 다시 시도하세요.";
      }
      return msg;
    }

    function esc(v){
      return String(v ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }

    async function api(url, opts){
      const res = await fetch(url, Object.assign({ credentials: "include" }, opts || {}));
      const txt = await res.text();
      let data = {};
      try { data = JSON.parse(txt); } catch (_) {}
      if (res.status === 401) {
        alert("주차관리 세션이 만료되었습니다. 다시 로그인해 주세요.");
        location.href = boot.back_url || "/pwa/";
        throw new Error("세션 만료");
      }
      if (res.status === 403) {
        throw new Error((data && data.detail) || "권한이 없습니다.");
      }
      if (!res.ok) throw new Error((data && data.detail) || txt || `HTTP ${res.status}`);
      return data;
    }

    function roleLabel(){
      const r = String(boot.role || "").trim().toLowerCase();
      if (r === "admin") return "관리자";
      if (r === "viewer") return "일반사용자";
      if (r === "guard") {
        const portal = String(boot.portal_permission_level || "").trim().toLowerCase();
        if (portal === "security_guard") return "보안/경비";
        if (portal === "site_admin") return "단지대표자";
        if (portal === "admin") return "관리자";
        return "단지대표자";
      }
      return r || "-";
    }

    function ctxLine(){
      const role = roleLabel();
      const site = [boot.site_code || "-", boot.site_name || ""].filter(Boolean).join(" / ");
      $("ctxLine").textContent = `단지: ${site} · 사용자: ${boot.display_name || "-"} · 권한: ${role}`;
      $("permLine").textContent = isManager
        ? "현재 세션은 주차관리 전체 메뉴(차량/위반/엑셀/스캐너)를 사용할 수 있습니다."
        : "현재 세션은 조회 메뉴만 사용할 수 있습니다.";
      $("btnBack").setAttribute("href", boot.back_url || "/pwa/");
      $("btnScanner").setAttribute("href", boot.scanner_url || "./scanner");
    }

    function setSystemLine(text, ok){
      const el = $("sysLine");
      if (!el) return;
      const msg = String(text || "").trim();
      el.textContent = msg;
      el.style.color = msg ? (ok ? "#8ee6a6" : "#ffb4b4") : "";
    }

    function fillForm(item){
      $("fPlate").value = item.plate || "";
      $("fStatus").value = item.status || "active";
      $("fUnit").value = item.unit || "";
      $("fOwner").value = item.owner_name || "";
      $("fFrom").value = item.valid_from || "";
      $("fTo").value = item.valid_to || "";
      $("fNote").value = item.note || "";
      state.selectedPlate = item.plate || "";
    }

    function clearForm(){
      fillForm({ status: "active" });
      state.selectedPlate = "";
      document.querySelectorAll("#tbVehicles tbody tr").forEach((tr)=>tr.classList.remove("sel"));
    }

    function payloadFromForm(){
      return {
        plate: ($("fPlate").value || "").trim(),
        status: ($("fStatus").value || "active").trim(),
        unit: ($("fUnit").value || "").trim() || null,
        owner_name: ($("fOwner").value || "").trim() || null,
        valid_from: ($("fFrom").value || "").trim() || null,
        valid_to: ($("fTo").value || "").trim() || null,
        note: ($("fNote").value || "").trim() || null,
      };
    }

    function renderVehicles(items){
      state.vehicles = items || [];
      const tbody = qs("#tbVehicles tbody");
      const wrap = $("vehicleTableWrap");
      if (!tbody) return;
      if (!items || !items.length){
        tbody.innerHTML = `<tr><td colspan="8">차량 데이터가 없습니다.</td></tr>`;
        if (wrap) {
          wrap.style.maxHeight = "none";
          wrap.style.overflowY = "visible";
        }
        return;
      }
      tbody.innerHTML = items.map((it) => `
        <tr data-plate="${esc(it.plate)}">
          <td>${esc(it.plate)}</td>
          <td>${esc(it.status_ko || it.status)}</td>
          <td>${esc(it.unit || "-")}</td>
          <td>${esc(it.owner_name || "-")}</td>
          <td>${esc(it.valid_from || "-")}</td>
          <td>${esc(it.valid_to || "-")}</td>
          <td>${esc(it.note || "-")}</td>
          <td>${esc(it.updated_at || "-")}</td>
        </tr>
      `).join("");

      if (wrap) {
        if (items.length > 15) {
          const headerRow = qs("#tbVehicles thead tr");
          const sampleRow = tbody.querySelector("tr[data-plate]");
          const headerHeight = headerRow ? Math.ceil(headerRow.getBoundingClientRect().height || 38) : 38;
          const rowHeight = sampleRow ? Math.ceil(sampleRow.getBoundingClientRect().height || 36) : 36;
          const maxHeight = headerHeight + (rowHeight * 15) + 2;
          wrap.style.maxHeight = `${maxHeight}px`;
          wrap.style.overflowY = "auto";
        } else {
          wrap.style.maxHeight = "none";
          wrap.style.overflowY = "visible";
        }
      }

      tbody.querySelectorAll("tr[data-plate]").forEach((tr) => {
        tr.addEventListener("click", () => {
          tbody.querySelectorAll("tr").forEach((x)=>x.classList.remove("sel"));
          tr.classList.add("sel");
          const p = tr.getAttribute("data-plate");
          const found = state.vehicles.find((v) => String(v.plate) === String(p));
          if (found) fillForm(found);
        });
      });
    }

    function renderViolations(items){
      const tbody = qs("#tbViolations tbody");
      const wrap = $("violationsTableWrap");
      if (!tbody) return;
      if (!items || !items.length){
        tbody.innerHTML = `<tr><td colspan="7">위반기록이 없습니다.</td></tr>`;
        if (wrap) {
          wrap.style.maxHeight = "none";
          wrap.style.overflowY = "visible";
        }
        return;
      }
      tbody.innerHTML = items.map((it) => `
        <tr>
          <td>${esc(it.created_at || "-")}</td>
          <td>${esc(it.plate || "-")}</td>
          <td>${esc(it.verdict_ko || it.verdict || "-")}</td>
          <td>${esc(it.location || "-")}</td>
          <td>${esc(it.memo || "-")}</td>
          <td>${esc(it.inspector || "-")}</td>
          <td>${it.photo_path ? `<a href="${esc(it.photo_path)}" target="_blank" rel="noopener">보기</a>` : "-"}</td>
        </tr>
      `).join("");

      if (wrap) {
        if (items.length > 18) {
          const headerRow = qs("#tbViolations thead tr");
          const sampleRow = tbody.querySelector("tr");
          const headerHeight = headerRow ? Math.ceil(headerRow.getBoundingClientRect().height || 38) : 38;
          const rowHeight = sampleRow ? Math.ceil(sampleRow.getBoundingClientRect().height || 36) : 36;
          const maxHeight = headerHeight + (rowHeight * 18) + 2;
          wrap.style.maxHeight = `${maxHeight}px`;
          wrap.style.overflowY = "auto";
        } else {
          wrap.style.maxHeight = "none";
          wrap.style.overflowY = "visible";
        }
      }
    }

    async function loadVehicles(){
      if (!isManager) return;
      const q = ($("qVehicle").value || "").trim();
      const u = new URL(apiUrl("vehicles"), location.origin);
      if (q) u.searchParams.set("q", q);
      u.searchParams.set("limit", "500");
      const data = await api(u.toString());
      renderVehicles(data.items || []);
      setMsg("msgVehicle", `차량 ${data.count || 0}건`, true);
    }

    async function loadViolations(){
      const u = new URL(apiUrl("violations/recent"), location.origin);
      u.searchParams.set("limit", "100");
      const data = await api(u.toString());
      const items = (data.items || []).map((x) => Object.assign({}, x, {
        verdict_ko: ({ OK:"정상등록", UNREGISTERED:"미등록", BLOCKED:"차단차량", EXPIRED:"기간만료", TEMP:"임시등록" }[x.verdict] || x.verdict || "-"),
      }));
      renderViolations(items);
    }

    async function loadSystemCheck(){
      if (!isManager) {
        setSystemLine("시스템 점검: 조회권한(일반사용자)", true);
        return;
      }
      try {
        const r = await api(apiUrl("system_check"));
        if (r && r.ok) {
          const pk = Array.isArray(r.vehicles_pk) ? r.vehicles_pk.join(", ") : "-";
          setSystemLine(`시스템 점검 정상 · DB쓰기:${r.can_write ? "OK" : "FAIL"} · PK:${pk} · site:${r.site_code}`, true);
        } else {
          setSystemLine(`시스템 점검 실패: ${(r && r.detail) || "unknown"}`, false);
        }
      } catch (e) {
        setSystemLine(`시스템 점검 실패: ${e.message || e}`, false);
      }
    }

    async function createVehicle(){
      const payload = payloadFromForm();
      const data = await api(apiUrl("vehicles"), {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      setMsg("msgVehicle", `신규 등록 완료: ${data.item.plate}`, true);
      await loadVehicles();
      clearForm();
    }

    async function updateVehicle(){
      const payload = payloadFromForm();
      const plate = state.selectedPlate || payload.plate;
      if (!plate) throw new Error("수정할 차량을 선택하세요.");
      const data = await api(apiUrl(`vehicles/${encodeURIComponent(plate)}`), {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      setMsg("msgVehicle", `수정 완료: ${data.item.plate}`, true);
      await loadVehicles();
    }

    async function deleteVehicle(){
      const plate = state.selectedPlate || ($("fPlate").value || "").trim();
      if (!plate) throw new Error("삭제할 차량을 선택하세요.");
      if (!confirm(`차량 ${plate} 를 삭제하시겠습니까?`)) return;
      await api(apiUrl(`vehicles/${encodeURIComponent(plate)}`), { method: "DELETE" });
      setMsg("msgVehicle", `삭제 완료: ${plate}`, true);
      await loadVehicles();
      clearForm();
    }

    async function importExcel(){
      const fileInput = $("excelFile");
      const file = fileInput && fileInput.files && fileInput.files[0];
      if (!file) throw new Error("엑셀(.xlsx) 파일을 선택하세요.");
      const fd = new FormData();
      fd.append("file", file);
      const data = await api(apiUrl("vehicles/import_excel"), { method: "POST", body: fd });
      const lines = [];
      lines.push(data.message || "가져오기 결과");
      lines.push(`신규등록: ${data.inserted || 0}건, 수정: ${data.updated || 0}건, 빈행건너뜀: ${data.skipped_empty || 0}건`);
      if (data.error_count) {
        lines.push(`형식오류: ${data.error_count}건`);
        (data.errors || []).slice(0, 5).forEach((e) => {
          lines.push(`- ${e.row}행: ${e.message}`);
        });
      }
      if ((data.unknown_headers || []).length) {
        lines.push(`인식되지 않은 컬럼: ${(data.unknown_headers || []).join(", ")}`);
      }
      if (data.guidance && typeof data.guidance === "object") {
        const g = data.guidance;
        if (g.status_guide) lines.push(`상태 안내: ${g.status_guide}`);
        if (g.date_guide) lines.push(`날짜 안내: ${g.date_guide}`);
      }
      setMsg("msgImport", lines.join("\n"), !data.error_count);
      await loadVehicles();
      fileInput.value = "";
    }

    function clearPortalAuthStore() {
      const tokenKey = "ka_part_auth_token_v1";
      const userKey = "ka_part_auth_user_v1";
      const logoutMarkKey = "ka_parking_manual_logout_v1";
      try { sessionStorage.setItem(logoutMarkKey, "1"); } catch (_e) {}
      try { sessionStorage.removeItem(tokenKey); } catch (_e) {}
      try { sessionStorage.removeItem(userKey); } catch (_e) {}
      try { localStorage.removeItem(tokenKey); } catch (_e) {}
      try { localStorage.removeItem(userKey); } catch (_e) {}
    }

    async function runIntegratedLogout() {
      try {
        await fetch(boot.logout_url || "./logout", { method: "POST", credentials: "include" });
      } catch (_) {}
      try {
        const portalLogoutUrl = String(boot.portal_logout_url || "/api/auth/logout").trim();
        if (portalLogoutUrl) {
          await fetch(portalLogoutUrl, { method: "POST", credentials: "include" });
        }
      } catch (_) {}
      clearPortalAuthStore();
      location.href = boot.logout_redirect_url || "/pwa/login.html";
    }

    function _isSmallScreen(){
      try { return !!(window.matchMedia && window.matchMedia("(max-width: 900px)").matches); } catch (_) { return false; }
    }

    function _setActiveTab(tabName){
      const t = String(tabName || "").trim().toLowerCase();
      const sVehicles = $("sectionVehicles");
      const sViolations = $("sectionViolations");
      if (sVehicles) sVehicles.classList.toggle("active", t === "vehicles");
      if (sViolations) sViolations.classList.toggle("active", t === "violations");
      const tabVehicles = $("tabVehicles");
      const tabViolations = $("tabViolations");
      if (tabVehicles) tabVehicles.classList.toggle("active", t === "vehicles");
      if (tabViolations) tabViolations.classList.toggle("active", t === "violations");
    }

    function enableTabbedUI(){
      if (!_isSmallScreen()) return;
      if (!$("bottomTabs")) return;

      const initial = isManager ? "vehicles" : "violations";
      _setActiveTab(initial);

      const tabScanner = $("tabScanner");
      if (tabScanner) tabScanner.setAttribute("href", boot.scanner_url || "./scanner");

      if (!isManager) {
        const tv = $("tabVehicles");
        if (tv) {
          tv.disabled = true;
          tv.classList.add("disabled");
        }
        const sv = $("sectionVehicles");
        if (sv) sv.style.display = "none";
      }

      const tabVehicles = $("tabVehicles");
      if (tabVehicles) {
        tabVehicles.addEventListener("click", () => {
          if (tabVehicles.disabled) return;
          _setActiveTab("vehicles");
          try { window.scrollTo({ top: 0, behavior: "smooth" }); } catch (_) { window.scrollTo(0, 0); }
        });
      }
      const tabViolations = $("tabViolations");
      if (tabViolations) {
        tabViolations.addEventListener("click", () => {
          _setActiveTab("violations");
          try { window.scrollTo({ top: 0, behavior: "smooth" }); } catch (_) { window.scrollTo(0, 0); }
        });
      }
      const tabLogout = $("tabLogout");
      if (tabLogout) {
        tabLogout.addEventListener("click", (e) => {
          e.preventDefault();
          runIntegratedLogout();
        });
      }

      try {
        const adv = $("advFields");
        if (adv && adv.tagName === "DETAILS") adv.open = false;
      } catch (_) {}

      document.body.classList.add("tabbed");
    }

    function bindEvents(){
      on("btnRefreshAll", "click", async () => {
        try {
          await loadVehicles();
          await loadViolations();
        } catch (e) {
          setMsg("msgVehicle", `새로고침 실패: ${e.message || e}`, false);
        }
      });

      on("btnReloadViolations", "click", async () => {
        try { await loadViolations(); } catch (e) {}
      });

      on("btnLogout", "click", async () => {
        await runIntegratedLogout();
      });

      if (!isManager) return;

      on("btnCreate", "click", async () => {
        try {
          await createVehicle();
        } catch (e) {
          setMsg("msgVehicle", `신규등록 실패: ${friendlyErrMsg(e)}`, false);
          if (String((e && e.message) || "").includes("이미 등록된 차량번호")) {
            try { await loadVehicles(); } catch (_) {}
          }
        }
      });
      on("btnUpdate", "click", async () => {
        try { await updateVehicle(); } catch (e) { setMsg("msgVehicle", `수정 실패: ${friendlyErrMsg(e)}`, false); }
      });
      on("btnDelete", "click", async () => {
        try { await deleteVehicle(); } catch (e) { setMsg("msgVehicle", `삭제 실패: ${friendlyErrMsg(e)}`, false); }
      });
      on("btnClear", "click", () => clearForm());
      on("btnImportExcel", "click", async () => {
        try { await importExcel(); } catch (e) { setMsg("msgImport", `엑셀 가져오기 실패: ${friendlyErrMsg(e)}`, false); }
      });
      const qVehicle = $("qVehicle");
      if (qVehicle) {
        qVehicle.addEventListener("keydown", async (e) => {
          if (e.key !== "Enter") return;
          try { await loadVehicles(); } catch (err) { setMsg("msgVehicle", `검색 실패: ${friendlyErrMsg(err)}`, false); }
        });
      }
    }

    async function init(){
      ctxLine();
      if (!isManager) {
        $("managerOnly").style.display = "none";
      } else {
        clearForm();
      }
      try { enableTabbedUI(); } catch (_e) {}
      bindEvents();
      window.__parkingAdminMainReady = true;
      await loadSystemCheck();
      if (isManager) {
        await loadVehicles();
      }
      await loadViolations();
    }

    init().catch((e) => {
      setMsg("msgVehicle", `초기화 실패: ${e.message || e}`, false);
    });
  })();
  </script>
  <script>
  (function(){
    function fallbackBind(){
      if (window.__parkingAdminMainReady) return;
      try { document.body.classList.remove("tabbed"); } catch (_e) {}
      const boot = window.__ADMIN_BOOT__ || {};
      const $ = (id) => document.getElementById(id);
      const msg = (text, ok) => {
        const el = $("msgVehicle");
        if (!el) return;
        el.textContent = text || "";
        el.className = "msg" + (text ? (ok ? " ok" : " err") : "");
      };
      const bind = (id, fn) => {
        const el = $(id);
        if (!el || el.dataset.fallbackBound === "1") return;
        el.dataset.fallbackBound = "1";
        el.addEventListener("click", fn);
      };
      const clearPortalAuthStore = () => {
        const tokenKey = "ka_part_auth_token_v1";
        const userKey = "ka_part_auth_user_v1";
        const logoutMarkKey = "ka_parking_manual_logout_v1";
        try { sessionStorage.setItem(logoutMarkKey, "1"); } catch (_e) {}
        try { sessionStorage.removeItem(tokenKey); } catch (_e) {}
        try { sessionStorage.removeItem(userKey); } catch (_e) {}
        try { localStorage.removeItem(tokenKey); } catch (_e) {}
        try { localStorage.removeItem(userKey); } catch (_e) {}
      };
      const apiBase = String(boot.api_base || "/parking/api/session").replace(/\/+$/, "");
      const api = async (path, opts) => {
        const p = String(path || "").replace(/^\/+/, "");
        const res = await fetch(`${apiBase}/${p}`, Object.assign({ credentials: "include" }, opts || {}));
        const txt = await res.text();
        let data = {};
        try { data = JSON.parse(txt); } catch (_) {}
        if (!res.ok) throw new Error((data && data.detail) || txt || `HTTP ${res.status}`);
        return data;
      };
      const formPayload = () => ({
        plate: ($("fPlate")?.value || "").trim(),
        status: ($("fStatus")?.value || "active").trim(),
        unit: ($("fUnit")?.value || "").trim() || null,
        owner_name: ($("fOwner")?.value || "").trim() || null,
        valid_from: ($("fFrom")?.value || "").trim() || null,
        valid_to: ($("fTo")?.value || "").trim() || null,
        note: ($("fNote")?.value || "").trim() || null,
      });
      const clear = () => {
        if ($("fPlate")) $("fPlate").value = "";
        if ($("fStatus")) $("fStatus").value = "active";
        if ($("fUnit")) $("fUnit").value = "";
        if ($("fOwner")) $("fOwner").value = "";
        if ($("fFrom")) $("fFrom").value = "";
        if ($("fTo")) $("fTo").value = "";
        if ($("fNote")) $("fNote").value = "";
      };

      bind("btnLogout", async () => {
        try {
          await fetch(boot.logout_url || "./logout", { method: "POST", credentials: "include" });
        } catch (_e) {}
        try {
          const portalLogoutUrl = String(boot.portal_logout_url || "/api/auth/logout").trim();
          if (portalLogoutUrl) {
            await fetch(portalLogoutUrl, { method: "POST", credentials: "include" });
          }
        } catch (_e) {}
        clearPortalAuthStore();
        location.href = boot.logout_redirect_url || "/pwa/login.html";
      });

      bind("btnCreate", async () => {
        try {
          const payload = formPayload();
          const out = await api("vehicles", {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload),
          });
          msg(`신규 등록 완료: ${(out && out.item && out.item.plate) || payload.plate}`, true);
        } catch (e) {
          msg(`신규등록 실패: ${e.message || e}`, false);
        }
      });

      bind("btnUpdate", async () => {
        try {
          const payload = formPayload();
          const plate = payload.plate;
          if (!plate) throw new Error("수정할 차량번호를 입력하세요.");
          const out = await api(`vehicles/${encodeURIComponent(plate)}`, {
            method: "PUT",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload),
          });
          msg(`수정 완료: ${(out && out.item && out.item.plate) || plate}`, true);
        } catch (e) {
          msg(`수정 실패: ${e.message || e}`, false);
        }
      });

      bind("btnDelete", async () => {
        try {
          const plate = ($("fPlate")?.value || "").trim();
          if (!plate) throw new Error("삭제할 차량번호를 입력하세요.");
          if (!confirm(`차량 ${plate} 를 삭제하시겠습니까?`)) return;
          await api(`vehicles/${encodeURIComponent(plate)}`, { method: "DELETE" });
          msg(`삭제 완료: ${plate}`, true);
          clear();
        } catch (e) {
          msg(`삭제 실패: ${e.message || e}`, false);
        }
      });

      bind("btnClear", () => {
        clear();
        msg("입력값을 초기화했습니다.", true);
      });

      bind("btnImportExcel", async () => {
        try {
          const fileInput = $("excelFile");
          const file = fileInput && fileInput.files && fileInput.files[0];
          if (!file) throw new Error("엑셀(.xlsx) 파일을 선택하세요.");
          const fd = new FormData();
          fd.append("file", file);
          const out = await api("vehicles/import_excel", { method: "POST", body: fd });
          msg(`엑셀 처리 완료: 신규 ${out.inserted || 0}건, 수정 ${out.updated || 0}건`, true);
        } catch (e) {
          msg(`엑셀 불러오기 실패: ${e.message || e}`, false);
        }
      });
    }

    setTimeout(fallbackBind, 700);
  })();
  </script>
</body>
</html>
"""


@app.get("/admin2", response_class=HTMLResponse)
def admin2(request: Request):
    s = read_session(request)
    if not s:
        if LOCAL_LOGIN_ENABLED:
            raise HTTPException(status_code=401, detail="Login required")
        return HTMLResponse(_auto_entry_page(app_url("/admin2")), status_code=200)

    if s.get("r") not in {"admin", "guard", "viewer"}:
        raise HTTPException(status_code=403, detail="Forbidden")

    role = str(s.get("r") or "").strip().lower()
    portal_permission_level = str(s.get("portal_permission_level") or "").strip().lower()
    site_code = _session_site_code(s)
    site_name = str(s.get("site_name") or "").strip()
    display_name = str(s.get("display_name") or s.get("u") or "").strip() or "사용자"
    logout_redirect_url = app_url("/login") if LOCAL_LOGIN_ENABLED else portal_login_url(app_url("/admin2"))
    context = {
        "role": role,
        "is_manager": _is_manager_session(s),
        "portal_permission_level": portal_permission_level,
        "site_code": site_code,
        "site_name": site_name,
        "display_name": display_name,
        "api_base": app_url("/api/session"),
        "scanner_url": app_url("/scanner"),
        "logout_url": app_url("/logout"),
        "portal_logout_url": "/api/auth/logout",
        "logout_redirect_url": logout_redirect_url,
        "back_url": "/pwa/",
    }
    html = ADMIN2_HTML_TEMPLATE.replace("__ADMIN_BOOT_JSON__", json.dumps(context, ensure_ascii=False))
    return HTMLResponse(html, status_code=200)


SCANNER_HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,viewport-fit=cover" />
  <meta name="theme-color" content="#0a2633" />
  <title>주차 스캐너</title>
  <link rel="manifest" href="./manifest.webmanifest" />
  <style>
    :root { --bg:#081923; --panel:#102533; --line:#244255; --text:#e7f1f6; --muted:#8cabbb; --ok:#16a34a; --bad:#dc2626; --warn:#d97706; --btn:#1d4ed8; }
    * { box-sizing: border-box; }
    body { margin:0; font-family: "Noto Sans KR", system-ui, sans-serif; background: radial-gradient(circle at top, #123347, #081923 58%); color:var(--text); }
    .app { max-width: 920px; margin: 0 auto; padding: 16px; padding: calc(16px + env(safe-area-inset-top)) calc(16px + env(safe-area-inset-right)) calc(16px + env(safe-area-inset-bottom)) calc(16px + env(safe-area-inset-left)); display: grid; gap: 12px; }
    .panel { background: rgba(16,37,51,.92); border: 1px solid var(--line); border-radius: 14px; padding: 14px; }
    .head { display:flex; align-items:center; justify-content:space-between; gap:10px; flex-wrap: wrap; }
    .title { margin: 0; font-size: 20px; }
    h2 { margin: 0; font-size: 18px; }
    .ctx { color: var(--muted); font-size: 13px; }
    .row { display:flex; gap:8px; margin-top: 8px; flex-wrap: wrap; }
    .btn { border:1px solid var(--line); background: var(--btn); color:white; border-radius:10px; padding:10px 12px; font-weight:700; cursor:pointer; text-decoration:none; display:inline-block; }
    .btn.ghost { background: transparent; color: var(--text); }
    .btn.warn { background: var(--warn); border-color: #b45309; }
    input { width: 100%; background:#0b1f2b; color:var(--text); border:1px solid var(--line); border-radius:10px; padding:10px 12px; }
    .camera-wrap { position:relative; margin-top:8px; border-radius:12px; overflow:hidden; border:1px solid var(--line); background:#031017; }
    video { width:100%; height:min(28vh, 190px); display:block; object-fit:cover; background:#031017; }
    .plate-guide { position:absolute; left:16%; top:52%; width:68%; height:30%; border:2px dashed rgba(111,190,245,.95); border-radius:10px; box-shadow:0 0 0 9999px rgba(3,16,23,.26); pointer-events:none; }
    .result { border-radius:12px; padding: 10px 12px; border:1px solid var(--line); background:#0a1f2b; }
    .result.ok { border-color: #166534; background: #052112; }
    .result.bad { border-color: #7f1d1d; background: #2d0b0b; }
    .result.warn { border-color: #92400e; background: #2c1508; }
    .hint { font-size: 12px; color: var(--muted); }
    .list { display:grid; gap:8px; margin-top:8px; }
    .item { border:1px solid var(--line); border-radius:10px; padding:8px 10px; background:#0a1f2b; }
    .item .meta { color:var(--muted); font-size:12px; margin-top:4px; }
    @media (max-width: 900px) {
      .title { font-size: 18px; }
      h2 { font-size: 16px; }
      input { font-size: 16px; }
    }
  </style>
</head>
<body>
  <main class="app">
    <section class="panel">
      <div class="head">
        <h1 class="title">주차 단속 스캐너</h1>
        <a href="./admin2" class="btn ghost">관리화면</a>
      </div>
      <p class="ctx" id="ctxLine">로딩 중...</p>
      <p class="hint">촬영 저장 없이 현재 카메라 프레임을 메모리에서만 읽어 즉시 분석합니다.</p>
      <div class="camera-wrap">
        <video id="video" playsinline autoplay muted></video>
        <div class="plate-guide" aria-hidden="true"></div>
      </div>
      <canvas id="canvas" hidden></canvas>
      <div class="row">
        <button id="btnCam" class="btn ghost">카메라 켜기</button>
        <button id="btnShot" class="btn">캡처+OCR(번호판영역)</button>
        <button id="btnAutoScan" class="btn ghost">즉시스캔 시작</button>
      </div>
      <div class="row">
        <input id="plateInput" placeholder="예: 123가4567" />
      </div>
      <div class="row">
        <button id="btnCheck" class="btn">불법주차 조회</button>
      </div>
      <p class="hint" id="hintLine">OCR 대기 중</p>
      <article id="result" class="result">
        <strong>조회 결과</strong>
        <div id="resultText">조회 전</div>
      </article>
    </section>

    <section class="panel">
      <h2 style="margin:0 0 8px 0;">위반 로그 저장</h2>
      <div class="row">
        <input id="locationInput" placeholder="위치(선택)" />
      </div>
      <div class="row">
        <input id="memoInput" placeholder="메모(선택)" />
      </div>
      <div class="row">
        <button id="btnSaveViolation" class="btn warn">현재 판정으로 위반기록 저장</button>
      </div>
      <div class="list" id="recentList"></div>
    </section>
  </main>

  <script>window.__BOOT__ = __BOOTSTRAP_JSON__;</script>
  <script src="https://cdn.jsdelivr.net/npm/tesseract.js@5/dist/tesseract.min.js" defer></script>
  <script>
  (function(){
    const $ = (id) => document.getElementById(id);
    const state = {
      stream: null,
      cameraStarting: null,
      last: null,
      autoScan: false,
      autoScanTimer: null,
      ocrBusy: false,
      lastAutoPlate: "",
      lastAutoAt: 0,
    };
    const ILLEGAL = new Set(["UNREGISTERED", "BLOCKED", "EXPIRED"]);
    const VERDICT_KO = { OK:"정상등록", UNREGISTERED:"미등록", BLOCKED:"차단차량", EXPIRED:"기간만료", TEMP:"임시등록" };
    const ROI_RECT = { x: 0.16, y: 0.52, w: 0.68, h: 0.30 };
    const AUTO_SCAN_INTERVAL_MS = 1400;
    const AUTO_SCAN_DEDUP_MS = 4500;
    const CAMERA_READY_TIMEOUT_MS = 7000;
    const OCR_READY_TIMEOUT_MS = 12000;
    const OCR_READY_POLL_MS = 250;
    const OCR_PLATE_WHITELIST = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ가나다라마바사아자차카타파하거너더러머버서어저처커터퍼허고노도로모보소오조초코토포호";
    const boot = window.__BOOT__ || {};

    const video = $("video");
    const canvas = $("canvas");
    const plateInput = $("plateInput");
    const hintLine = $("hintLine");
    const result = $("result");
    const resultText = $("resultText");
    const recentList = $("recentList");

    function setCtx() {
      const site = [boot.site_code || "-", boot.site_name || ""].filter(Boolean).join(" / ");
      const user = boot.display_name || boot.user_login || "-";
      $("ctxLine").textContent = `단지: ${site} · 사용자: ${user}`;
    }

    function setHint(v) { hintLine.textContent = v; }

    function sleep(ms) {
      return new Promise((resolve) => setTimeout(resolve, Number(ms) || 0));
    }

    function hasLiveStream(stream) {
      if (!stream) return false;
      const tracks = typeof stream.getVideoTracks === "function" ? stream.getVideoTracks() : [];
      if (!tracks.length) return false;
      return tracks.some((t) => t && t.readyState === "live");
    }

    function setCameraButtonState(on) {
      const btn = $("btnCam");
      if (!btn) return;
      if (on) {
        btn.textContent = "카메라 끄기";
        btn.classList.add("warn");
      } else {
        btn.textContent = "카메라 켜기";
        btn.classList.remove("warn");
      }
    }

    function describeCameraError(err) {
      const name = String(err && err.name || "").trim();
      if (name === "NotAllowedError" || name === "PermissionDeniedError") {
        return "카메라 권한이 거부되었습니다. 브라우저/OS 설정에서 카메라 권한을 허용하세요.";
      }
      if (name === "NotFoundError" || name === "DevicesNotFoundError") {
        return "사용 가능한 카메라 장치를 찾지 못했습니다.";
      }
      if (name === "NotReadableError" || name === "TrackStartError") {
        return "카메라가 다른 앱에서 사용 중이거나 장치에 접근할 수 없습니다.";
      }
      if (name === "OverconstrainedError" || name === "ConstraintNotSatisfiedError") {
        return "카메라 조건이 맞지 않습니다. 기본 카메라로 다시 시도하세요.";
      }
      if (name === "SecurityError") {
        return "보안 정책으로 카메라를 사용할 수 없습니다. HTTPS 접속인지 확인하세요.";
      }
      const msg = String(err && err.message || "").trim();
      return msg || "카메라 시작 중 알 수 없는 오류가 발생했습니다.";
    }

    async function waitForVideoReady(timeoutMs = CAMERA_READY_TIMEOUT_MS) {
      const timeout = Math.max(500, Number(timeoutMs) || CAMERA_READY_TIMEOUT_MS);
      const startAt = Date.now();
      while ((Date.now() - startAt) < timeout) {
        if ((video.readyState || 0) >= 2 && Number(video.videoWidth || 0) > 1 && Number(video.videoHeight || 0) > 1) {
          return true;
        }
        await sleep(80);
      }
      throw new Error("카메라 프레임 준비 시간이 초과되었습니다.");
    }

    async function ensureOcrReady(timeoutMs = OCR_READY_TIMEOUT_MS) {
      if (window.Tesseract && typeof window.Tesseract.recognize === "function") return true;
      const timeout = Math.max(1000, Number(timeoutMs) || OCR_READY_TIMEOUT_MS);
      const startAt = Date.now();
      while ((Date.now() - startAt) < timeout) {
        await sleep(OCR_READY_POLL_MS);
        if (window.Tesseract && typeof window.Tesseract.recognize === "function") return true;
      }
      return false;
    }

    function stopCamera({ clearHint = false } = {}) {
      if (state.stream) {
        try {
          state.stream.getTracks().forEach((t) => t.stop());
        } catch (_e) {}
      }
      state.stream = null;
      state.cameraStarting = null;
      if (video) video.srcObject = null;
      setCameraButtonState(false);
      if (clearHint) setHint("카메라가 꺼졌습니다.");
    }

    function norm(raw) {
      return String(raw || "").replace(/[^0-9A-Za-z가-힣]/g, "").toUpperCase().trim();
    }

    function findPlate(raw) {
      const txt = norm(raw);
      if (!txt) return "";
      const ko = txt.match(/\d{2,3}[가-힣]\d{4}/g);
      if (ko && ko.length) return ko[0];
      const latin = txt.match(/\d{2,3}[A-Z]\d{4}/g);
      if (latin && latin.length) return latin[0];
      return "";
    }

    function setResult(kind, text) {
      result.className = "result" + (kind ? ` ${kind}` : "");
      resultText.textContent = text;
    }

    async function startCamera({ forceRestart = false } = {}) {
      if (!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia) {
        throw new Error("이 브라우저는 카메라 API를 지원하지 않습니다.");
      }

      if (state.cameraStarting) {
        return await state.cameraStarting;
      }

      if (hasLiveStream(state.stream) && !forceRestart) {
        if (video.srcObject !== state.stream) {
          video.srcObject = state.stream;
        }
        try {
          await video.play();
        } catch (_e) {}
        await waitForVideoReady(CAMERA_READY_TIMEOUT_MS);
        setCameraButtonState(true);
        setHint("카메라 활성화");
        return true;
      }

      if (state.stream && !hasLiveStream(state.stream)) {
        stopCamera();
      }

      const run = (async () => {
        const candidates = [
          { video: { facingMode: { ideal: "environment" }, width: { ideal: 1280 }, height: { ideal: 720 } }, audio: false },
          { video: { facingMode: "environment" }, audio: false },
          { video: true, audio: false },
        ];
        let lastErr = null;
        let nextStream = null;
        for (const constraints of candidates) {
          try {
            nextStream = await navigator.mediaDevices.getUserMedia(constraints);
            if (nextStream) break;
          } catch (e) {
            lastErr = e;
          }
        }
        if (!nextStream) {
          throw new Error(describeCameraError(lastErr));
        }

        state.stream = nextStream;
        video.srcObject = nextStream;
        video.setAttribute("playsinline", "true");
        video.muted = true;
        try {
          await video.play();
        } catch (_e) {}
        await waitForVideoReady(CAMERA_READY_TIMEOUT_MS);
        setCameraButtonState(true);
        setHint("카메라 활성화");
        return true;
      })();

      state.cameraStarting = run;
      try {
        return await run;
      } catch (e) {
        stopCamera();
        throw e;
      } finally {
        state.cameraStarting = null;
      }
    }

    function _clamp(v, min, max) {
      return Math.max(min, Math.min(max, v));
    }

    function capturePlateRoiCanvas() {
      const w = Number(video.videoWidth || 0);
      const h = Number(video.videoHeight || 0);
      if ((video.readyState || 0) < 2 || w < 2 || h < 2) {
        throw new Error("카메라 프레임이 아직 준비되지 않았습니다.");
      }

      canvas.width = w;
      canvas.height = h;
      const ctx = canvas.getContext("2d");
      try {
        ctx.drawImage(video, 0, 0, w, h);
      } catch (_e) {
        throw new Error("카메라 프레임 캡처에 실패했습니다. 잠시 후 다시 시도하세요.");
      }

      const rx = _clamp(Math.floor(w * ROI_RECT.x), 0, w - 2);
      const ry = _clamp(Math.floor(h * ROI_RECT.y), 0, h - 2);
      const rw = _clamp(Math.floor(w * ROI_RECT.w), 2, w - rx);
      const rh = _clamp(Math.floor(h * ROI_RECT.h), 2, h - ry);

      const roi = document.createElement("canvas");
      roi.width = rw * 2;
      roi.height = rh * 2;
      const rctx = roi.getContext("2d");
      rctx.imageSmoothingEnabled = false;
      rctx.drawImage(canvas, rx, ry, rw, rh, 0, 0, roi.width, roi.height);
      return roi;
    }

    function buildOcrVariant(srcCanvas, { invert = false } = {}) {
      const out = document.createElement("canvas");
      out.width = srcCanvas.width;
      out.height = srcCanvas.height;
      const ctx = out.getContext("2d", { willReadFrequently: true });
      ctx.drawImage(srcCanvas, 0, 0);
      const img = ctx.getImageData(0, 0, out.width, out.height);
      const data = img.data;
      const pxCount = data.length / 4;
      let sumGray = 0;

      for (let i = 0; i < data.length; i += 4) {
        const gray = (data[i] * 299 + data[i + 1] * 587 + data[i + 2] * 114) / 1000;
        sumGray += gray;
      }

      const avg = sumGray / Math.max(1, pxCount);
      for (let i = 0; i < data.length; i += 4) {
        const gray = (data[i] * 299 + data[i + 1] * 587 + data[i + 2] * 114) / 1000;
        const contrast = _clamp((gray - avg) * 1.85 + 128, 0, 255);
        let bw = contrast > 132 ? 255 : 0;
        if (invert) bw = 255 - bw;
        data[i] = bw;
        data[i + 1] = bw;
        data[i + 2] = bw;
      }
      ctx.putImageData(img, 0, 0);
      return out;
    }

    async function recognizePlateFromCanvas(targetCanvas, passName) {
      const out = await window.Tesseract.recognize(targetCanvas, "kor+eng", {
        tessedit_pageseg_mode: "7",
        preserve_interword_spaces: "0",
        tessedit_char_whitelist: OCR_PLATE_WHITELIST,
        user_defined_dpi: "300",
      });
      const text = String((out && out.data && out.data.text) || "");
      const plate = findPlate(text);
      const confidence = Number((out && out.data && out.data.confidence) || 0);
      return { plate, confidence, pass: passName };
    }

    async function runOcr() {
      const ocrReady = await ensureOcrReady();
      if (!ocrReady) {
        throw new Error("OCR 엔진 로딩이 지연되고 있습니다. 네트워크 상태를 확인한 뒤 다시 시도하세요.");
      }
      if (state.ocrBusy) return "";
      state.ocrBusy = true;
      try {
        setHint("번호판 영역을 즉시 분석 중...");
        const roi = capturePlateRoiCanvas();
        const variants = [
          { pass: "raw", canvas: roi },
          { pass: "bw", canvas: buildOcrVariant(roi, { invert: false }) },
          { pass: "inv", canvas: buildOcrVariant(roi, { invert: true }) },
        ];
        let bestPlate = "";
        let bestConfidence = -1;
        let bestPass = "";

        for (const variant of variants) {
          const r = await recognizePlateFromCanvas(variant.canvas, variant.pass);
          if (!r.plate) continue;
          if (r.confidence > bestConfidence) {
            bestPlate = r.plate;
            bestConfidence = r.confidence;
            bestPass = r.pass;
          }
          if (r.confidence >= 70) break;
        }

        if (!bestPlate) {
          setHint("번호판을 찾지 못했습니다. 가이드 박스에 번호판을 맞춰주세요.");
          return "";
        }
        setHint(`OCR 추출: ${bestPlate} (${bestPass})`);
        return bestPlate;
      } finally {
        state.ocrBusy = false;
      }
    }

    async function checkPlate(raw, source) {
      const plate = norm(raw);
      if (plate.length < 7) {
        setResult("warn", "번호판 입력값이 너무 짧습니다.");
        return;
      }
      const u = new URL("./api/session/plates/check", location.href);
      u.searchParams.set("plate", plate);
      const res = await fetch(u.toString(), { credentials: "include" });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.detail || `HTTP ${res.status}`);
      state.last = data;
      plateInput.value = data.plate || plate;
      const suffix = source ? ` [${source}]` : "";
      const verdictKo = VERDICT_KO[data.verdict] || data.verdict || "-";
      if (ILLEGAL.has(data.verdict)) {
        setResult("bad", `불법주차 의심 차량 (${verdictKo})${suffix} - ${data.message}`);
      } else if (data.verdict === "OK" || data.verdict === "TEMP") {
        setResult("ok", `정상/임시 등록 차량 (${verdictKo})${suffix} - ${data.message}`);
      } else {
        setResult("warn", `판정: ${verdictKo}${suffix}`);
      }
      await loadRecent();
    }

    async function analyzeAndCheck(source) {
      await startCamera();
      const plate = await runOcr();
      if (!plate) return false;

      const now = Date.now();
      if (source === "AUTO") {
        if (state.lastAutoPlate === plate && (now - state.lastAutoAt) < AUTO_SCAN_DEDUP_MS) {
          return false;
        }
        state.lastAutoPlate = plate;
        state.lastAutoAt = now;
      }

      plateInput.value = plate;
      await checkPlate(plate, source);
      return true;
    }

    async function saveViolation() {
      if (!state.last) {
        setResult("warn", "먼저 조회를 실행하세요.");
        return;
      }
      if (!ILLEGAL.has(state.last.verdict)) {
        setResult("warn", "현재 판정은 위반 저장 대상이 아닙니다.");
        return;
      }
      const payload = {
        plate: state.last.plate,
        verdict: state.last.verdict,
        location: ($("locationInput").value || "").trim() || null,
        memo: ($("memoInput").value || "").trim() || null,
      };
      const res = await fetch("./api/session/violations", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
        credentials: "include",
      });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.detail || `HTTP ${res.status}`);
      const verdictKo = VERDICT_KO[data.verdict] || data.verdict || "-";
      setResult("bad", `위반기록 저장 완료: ${data.plate} (${verdictKo})`);
      await loadRecent();
    }

    function renderRecent(items) {
      if (!items || !items.length) {
        recentList.innerHTML = '<div class="item">최근 위반기록 없음</div>';
        return;
      }
      recentList.innerHTML = items.map((it) => (
        `<div class="item"><strong>${it.plate || "-"}</strong> · ${it.verdict || "-"}`
        + `<div class="meta">${it.created_at || "-"} · ${it.inspector || "-"}</div></div>`
      )).join("");
    }

    async function loadRecent() {
      const u = new URL("./api/session/violations/recent", location.href);
      u.searchParams.set("limit", "20");
      const res = await fetch(u.toString(), { credentials: "include" });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.detail || `HTTP ${res.status}`);
      renderRecent(data.items || []);
    }

    function setAutoScanMode(on) {
      state.autoScan = !!on;
      const btn = $("btnAutoScan");
      if (!btn) return;
      btn.textContent = state.autoScan ? "즉시스캔 중지" : "즉시스캔 시작";
      if (state.autoScan) btn.classList.add("warn");
      else btn.classList.remove("warn");
    }

    function stopAutoScan() {
      if (state.autoScanTimer) {
        clearTimeout(state.autoScanTimer);
        state.autoScanTimer = null;
      }
      setAutoScanMode(false);
    }

    async function autoScanTick() {
      if (!state.autoScan) return;
      try {
        if (!hasLiveStream(state.stream)) {
          await startCamera();
        }
        await analyzeAndCheck("AUTO");
      } catch (e) {
        const msg = String(e && e.message ? e.message : e);
        setHint(`즉시스캔 오류: ${msg}`);
        if (msg.includes("권한") || msg.includes("지원하지") || msg.includes("보안")) {
          stopAutoScan();
        }
      } finally {
        if (state.autoScan) {
          state.autoScanTimer = setTimeout(autoScanTick, AUTO_SCAN_INTERVAL_MS);
        }
      }
    }

    $("btnCam").addEventListener("click", async () => {
      try {
        if (hasLiveStream(state.stream)) {
          stopAutoScan();
          stopCamera({ clearHint: true });
          return;
        }
        await startCamera({ forceRestart: true });
      } catch (e) {
        setHint(String(e.message || e));
      }
    });

    $("btnShot").addEventListener("click", async () => {
      try {
        await analyzeAndCheck("OCR");
      } catch (e) {
        setHint(`OCR 오류: ${String(e.message || e)}`);
      }
    });

    $("btnAutoScan").addEventListener("click", async () => {
      try {
        if (state.autoScan) {
          stopAutoScan();
          setHint("즉시스캔을 중지했습니다.");
          return;
        }
        await startCamera();
        if (state.autoScanTimer) {
          clearTimeout(state.autoScanTimer);
          state.autoScanTimer = null;
        }
        setAutoScanMode(true);
        setHint("즉시스캔 시작: 촬영 저장 없이 현재 프레임을 반복 분석합니다.");
        await autoScanTick();
      } catch (e) {
        stopAutoScan();
        setHint(`즉시스캔 시작 실패: ${String(e.message || e)}`);
      }
    });

    $("btnCheck").addEventListener("click", async () => {
      try {
        await checkPlate(plateInput.value, "MANUAL");
      } catch (e) {
        setResult("warn", `조회 실패: ${String(e.message || e)}`);
      }
    });

    $("btnSaveViolation").addEventListener("click", async () => {
      try {
        await saveViolation();
      } catch (e) {
        setResult("warn", `저장 실패: ${String(e.message || e)}`);
      }
    });

    window.addEventListener("beforeunload", () => {
      stopAutoScan();
      stopCamera();
    });

    document.addEventListener("visibilitychange", () => {
      if (!document.hidden) return;
      stopAutoScan();
      stopCamera();
    });

    if ("serviceWorker" in navigator) {
      navigator.serviceWorker.register("./sw.js").catch(() => {});
    }
    setCtx();
    loadRecent().catch(() => {});
  })();
  </script>
</body>
</html>
"""


@app.get("/scanner", response_class=HTMLResponse)
def scanner_page(request: Request):
    sess = read_session(request)
    if not sess:
        if LOCAL_LOGIN_ENABLED:
            raise HTTPException(status_code=401, detail="Login required")
        return HTMLResponse(_auto_entry_page(app_url("/scanner")), status_code=200)

    if sess.get("r") not in {"admin", "guard", "viewer"}:
        raise HTTPException(status_code=403, detail="Forbidden")

    context = {
        "site_code": _session_site_code(sess),
        "site_name": str(sess.get("site_name") or "").strip(),
        "display_name": str(sess.get("display_name") or "").strip(),
        "user_login": str(sess.get("u") or "").strip(),
    }
    html = SCANNER_HTML_TEMPLATE.replace("__BOOTSTRAP_JSON__", json.dumps(context, ensure_ascii=False))
    return HTMLResponse(html, status_code=200)


@app.get("/manifest.webmanifest")
def scanner_manifest():
    return JSONResponse(
        {
            "name": "주차 단속 스캐너",
            "short_name": "주차스캐너",
            "start_url": app_url("/scanner"),
            "scope": app_url("/"),
            "display": "standalone",
            "background_color": "#081923",
            "theme_color": "#0a2633",
            "description": "카메라로 번호판을 인식하고 불법주차 여부를 확인하는 앱",
        }
    )


@app.get("/sw.js")
def scanner_sw():
    sw = """
self.addEventListener('install', (event) => {
  event.waitUntil(self.skipWaiting());
});
self.addEventListener('activate', (event) => {
  event.waitUntil(self.clients.claim());
});
"""
    return Response(content=sw, media_type="application/javascript")
