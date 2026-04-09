from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _as_text(value: Any) -> str:
    return str(value or "").strip()


def build_ops_document_ledger_xlsx(
    *,
    tenant_label: str,
    selected_category: str,
    documents: Iterable[Dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "문서관리대장"

    title = "행정문서 관리대장"
    category_label = _as_text(selected_category) or "전체"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sheet["A1"] = title
    sheet["A2"] = f"사업장: {tenant_label}"
    sheet["A3"] = f"분류: {category_label}"
    sheet["F2"] = f"출력일시: {generated_at}"

    headers = [
        "제목",
        "분류",
        "상태",
        "담당",
        "기한",
        "문서번호",
        "대상/설비",
        "업체/상대처",
        "금액(원)",
        "기준일",
        "시작일",
        "종료일",
        "요약",
        "등록자",
        "등록일",
        "수정일",
    ]
    sheet.append([])
    sheet.append(headers)

    title_fill = PatternFill(fill_type="solid", fgColor="0D6A67")
    header_fill = PatternFill(fill_type="solid", fgColor="DDEEEB")

    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = title_fill
    for cell in sheet[5]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows: List[Dict[str, Any]] = list(documents)
    for item in rows:
        sheet.append(
            [
                _as_text(item.get("title")),
                _as_text(item.get("category")),
                _as_text(item.get("status")),
                _as_text(item.get("owner")),
                _as_text(item.get("due_date")),
                _as_text(item.get("reference_no")),
                _as_text(item.get("target_label")),
                _as_text(item.get("vendor_name")),
                _as_text(item.get("amount_total")),
                _as_text(item.get("basis_date")),
                _as_text(item.get("period_start")),
                _as_text(item.get("period_end")),
                _as_text(item.get("summary")),
                _as_text(item.get("created_by_label")),
                _as_text(item.get("created_at")),
                _as_text(item.get("updated_at")),
            ]
        )

    sheet.freeze_panes = "A6"
    column_widths = {
        "A": 30,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 42,
        "N": 16,
        "O": 20,
        "P": 20,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
