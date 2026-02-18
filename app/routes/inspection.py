from __future__ import annotations

import hashlib
import json
import os
import re
import secrets
import uuid
from io import BytesIO
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple

from fastapi import APIRouter, Body, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from ..db import _connect, get_auth_user_by_token, now_iso

router = APIRouter()

AUTH_COOKIE_NAME = (os.getenv("KA_AUTH_COOKIE_NAME") or "ka_part_auth_token").strip()
ALLOW_QUERY_ACCESS_TOKEN = (os.getenv("KA_ALLOW_QUERY_ACCESS_TOKEN") or "").strip().lower() in {"1", "true", "yes", "on"}
SECURITY_ROLE_KEYWORDS = ("보안", "경비")
PUBLIC_ACCESS_LOGIN_ID = ((os.getenv("KA_PUBLIC_FULL_ACCESS_LOGIN_ID") or "public_guest").strip().lower() or "public_guest")
RUN_CREATOR_ROLE_SET = {"최고/운영관리자", "최고관리자", "운영관리자", "단지대표자", "단지관리자", "사용자"}

ROOT_DIR = Path(__file__).resolve().parents[2]
UPLOAD_ROOT = (ROOT_DIR / "uploads" / "inspection").resolve()
PHOTO_ROOT = (UPLOAD_ROOT / "photos").resolve()
ARCHIVE_ROOT = (UPLOAD_ROOT / "archives").resolve()
PHOTO_MAX_FILE_BYTES = int(os.getenv("KA_INSPECTION_PHOTO_MAX_BYTES") or str(8 * 1024 * 1024))
APPROVAL_CODE_TTL_MINUTES = max(1, int(str(os.getenv("KA_INSPECTION_APPROVAL_CODE_TTL_MINUTES") or "10").strip() or "10"))
APPROVAL_CODE_DIGITS = max(4, min(8, int(str(os.getenv("KA_INSPECTION_APPROVAL_CODE_DIGITS") or "6").strip() or "6")))
APPROVAL_CODE_SECRET = str(os.getenv("KA_INSPECTION_APPROVAL_CODE_SECRET") or "inspection-approval-code").strip() or "inspection-approval-code"


class TargetPayload(BaseModel):
    site_code: str = ""
    name: str = Field(..., min_length=2, max_length=120)
    location: str = ""
    description: str = ""
    is_active: bool = True
    force_new: bool = False


class TargetUpdatePayload(BaseModel):
    name: str | None = None
    location: str | None = None
    description: str | None = None
    is_active: bool | None = None


class TemplateItemPayload(BaseModel):
    item_key: str = Field(..., min_length=1, max_length=80)
    item_text: str = Field(..., min_length=1, max_length=300)
    category: str = ""
    severity: int = Field(default=1, ge=1, le=3)
    sort_order: int = Field(default=0, ge=0, le=100000)
    requires_photo: bool = False
    requires_note: bool = False
    is_active: bool = True


class TemplatePayload(BaseModel):
    site_code: str = ""
    target_id: int = Field(..., ge=1)
    name: str = Field(..., min_length=2, max_length=160)
    period: str = Field(default="MONTHLY", max_length=20)
    is_active: bool = True
    force_new: bool = False
    auto_backup: bool = False
    items: List[TemplateItemPayload] = Field(default_factory=list)


class TemplateUpdatePayload(BaseModel):
    target_id: int | None = None
    name: str | None = None
    period: str | None = None
    is_active: bool | None = None
    auto_backup: bool = False
    items: List[TemplateItemPayload] | None = None


class RunCreatePayload(BaseModel):
    site_code: str = ""
    target_id: int = Field(..., ge=1)
    template_id: int = Field(..., ge=1)
    run_date: str = ""
    note: str = ""


class RunItemPatchPayload(BaseModel):
    id: int = Field(..., ge=1)
    result: str = Field(..., min_length=2, max_length=20)
    note: str = ""


class RunItemsPatchRequest(BaseModel):
    items: List[RunItemPatchPayload] = Field(default_factory=list)


class DecisionPayload(BaseModel):
    comment: str = ""
    approval_code: str = ""


def _now_ts() -> str:
    return now_iso()


def _clean_site_code(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if not raw:
        return ""
    return re.sub(r"[\s-]+", "", raw)


def _safe_date(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return datetime.now().strftime("%Y-%m-%d")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        raise HTTPException(status_code=400, detail="날짜 형식은 YYYY-MM-DD 입니다.")
    return raw


def _safe_period(value: Any) -> str:
    raw = str(value or "").strip().upper() or "MONTHLY"
    if raw not in {"DAILY", "WEEKLY", "MONTHLY", "QUARTERLY", "YEARLY"}:
        raise HTTPException(status_code=400, detail="period must be DAILY/WEEKLY/MONTHLY/QUARTERLY/YEARLY")
    return raw


def _safe_result(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if raw not in {"COMPLIANT", "NONCOMPLIANT", "NA"}:
        raise HTTPException(status_code=400, detail="result must be COMPLIANT/NONCOMPLIANT/NA")
    return raw


def _result_label(value: Any) -> str:
    code = _safe_result(value)
    if code == "COMPLIANT":
        return "양호"
    if code == "NONCOMPLIANT":
        return "부적합"
    return "해당없음"


def _normalize_approval_code(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or "").strip())


def _issue_random_approval_code() -> str:
    minimum = 10 ** (APPROVAL_CODE_DIGITS - 1)
    maximum = (10 ** APPROVAL_CODE_DIGITS) - 1
    return str(secrets.randbelow(maximum - minimum + 1) + minimum)


def _approval_code_hash(*, run_id: int, step_no: int, actor_login: str, code: str) -> str:
    raw = f"{int(run_id)}|{int(step_no)}|{str(actor_login or '').strip().lower()}|{_normalize_approval_code(code)}|{APPROVAL_CODE_SECRET}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _normalize_role_text(value: Any) -> str:
    return str(value or "").strip()


def _is_security_role(value: Any) -> bool:
    role = _normalize_role_text(value)
    if not role:
        return False
    compact = role.replace(" ", "")
    if compact == "보안/경비":
        return True
    return any(token in role for token in SECURITY_ROLE_KEYWORDS)


def _is_resident_or_board_role(value: Any) -> bool:
    role = _normalize_role_text(value)
    if role in {"입주민", "주민", "세대주민"}:
        return True
    return role in {"입대의", "입주자대표", "입주자대표회의"}


def _is_public_access_user(user: Dict[str, Any]) -> bool:
    login_id = str(user.get("login_id") or "").strip().lower()
    return bool(login_id) and login_id == PUBLIC_ACCESS_LOGIN_ID


def _is_admin(user: Dict[str, Any]) -> bool:
    return int(user.get("is_admin") or 0) == 1


def _is_site_admin(user: Dict[str, Any]) -> bool:
    return int(user.get("is_site_admin") or 0) == 1


def _is_run_creator(user: Dict[str, Any]) -> bool:
    if _is_admin(user) or _is_site_admin(user):
        return True
    role = _normalize_role_text(user.get("role"))
    return role in RUN_CREATOR_ROLE_SET


def _extract_access_token(request: Request) -> str:
    auth = (request.headers.get("Authorization") or "").strip()
    if auth.lower().startswith("bearer "):
        token = auth[7:].strip()
        if token:
            return token
    cookie_token = (request.cookies.get(AUTH_COOKIE_NAME) or "").strip()
    if cookie_token:
        return cookie_token
    if ALLOW_QUERY_ACCESS_TOKEN:
        token = (request.query_params.get("access_token") or "").strip()
        if token:
            return token
    raise HTTPException(status_code=401, detail="로그인이 필요합니다.")


def _require_auth(request: Request) -> Tuple[Dict[str, Any], str]:
    token = _extract_access_token(request)
    user = get_auth_user_by_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다. (세션 만료)")
    return user, token


def _require_inspection_user(request: Request) -> Tuple[Dict[str, Any], str]:
    user, token = _require_auth(request)
    if _is_public_access_user(user):
        raise HTTPException(status_code=403, detail="로그인 없이 사용자는 점검 모듈을 사용할 수 없습니다. 신규가입 후 사용해 주세요.")
    if _is_security_role(user.get("role")):
        raise HTTPException(status_code=403, detail="보안/경비 계정은 주차관리 모듈만 사용할 수 있습니다.")
    if _is_resident_or_board_role(user.get("role")):
        raise HTTPException(status_code=403, detail="입주민/입대의 계정은 점검 모듈을 사용할 수 없습니다.")
    return user, token


def _require_manager_user(request: Request) -> Tuple[Dict[str, Any], str]:
    user, token = _require_inspection_user(request)
    if not (_is_admin(user) or _is_site_admin(user)):
        raise HTTPException(status_code=403, detail="점검 모듈 관리자 권한이 필요합니다.")
    return user, token


def _require_run_creator_user(request: Request) -> Tuple[Dict[str, Any], str]:
    user, token = _require_inspection_user(request)
    if not _is_run_creator(user):
        raise HTTPException(status_code=403, detail="점검 생성 권한이 없습니다.")
    return user, token


def _scope_site_code(user: Dict[str, Any], requested_site_code: Any = "") -> str:
    requested = _clean_site_code(requested_site_code)
    user_site = _clean_site_code(user.get("site_code"))
    if _is_admin(user):
        if requested:
            return requested
        if user_site:
            return user_site
        raise HTTPException(status_code=400, detail="site_code가 필요합니다.")
    if not user_site:
        raise HTTPException(status_code=403, detail="계정에 site_code가 없습니다. 관리자에게 문의하세요.")
    if requested and requested != user_site:
        raise HTTPException(status_code=403, detail="소속 단지(site_code) 데이터만 접근할 수 있습니다.")
    return user_site


def _run_scope_check(user: Dict[str, Any], run_row: Dict[str, Any]) -> None:
    run_site = _clean_site_code(run_row.get("site_code"))
    scoped = _scope_site_code(user, run_site)
    if run_site != scoped:
        raise HTTPException(status_code=403, detail="소속 단지(site_code) 데이터만 접근할 수 있습니다.")
    if _is_admin(user) or _is_site_admin(user):
        return
    actor_login = str(user.get("login_id") or "").strip().lower()
    inspector_login = str(run_row.get("inspector_login") or "").strip().lower()
    if actor_login != inspector_login:
        raise HTTPException(status_code=403, detail="본인 점검만 접근할 수 있습니다.")


def _can_edit_run(user: Dict[str, Any], run_row: Dict[str, Any]) -> bool:
    if _is_admin(user) or _is_site_admin(user):
        return True
    actor_login = str(user.get("login_id") or "").strip().lower()
    inspector_login = str(run_row.get("inspector_login") or "").strip().lower()
    return actor_login == inspector_login


def _next_run_code(con, run_date: str) -> str:
    year = str(run_date or "").strip()[:4]
    if not re.fullmatch(r"\d{4}", year):
        year = datetime.now().strftime("%Y")
    row = con.execute(
        "SELECT run_code FROM inspection_runs WHERE run_code LIKE ? ORDER BY id DESC LIMIT 1",
        (f"INSP-{year}-%",),
    ).fetchone()
    if not row:
        return f"INSP-{year}-000001"
    raw = str(row["run_code"] or "")
    try:
        seq = int(raw.rsplit("-", 1)[-1]) + 1
    except Exception:
        seq = 1
    return f"INSP-{year}-{seq:06d}"


def _next_unique_master_name(con, table_name: str, site_code: str, base_name: str) -> str:
    table = str(table_name or "").strip().lower()
    if table not in {"inspection_targets", "inspection_templates"}:
        raise HTTPException(status_code=500, detail="internal table validation failed")
    clean_site = _clean_site_code(site_code)
    seed = str(base_name or "").strip() or ("점검대상" if table == "inspection_targets" else "점검표")

    row = con.execute(
        f"SELECT id FROM {table} WHERE site_code=? AND name=? LIMIT 1",
        (clean_site, seed),
    ).fetchone()
    if not row:
        return seed

    for i in range(2, 10000):
        candidate = f"{seed} ({i})"
        dup = con.execute(
            f"SELECT id FROM {table} WHERE site_code=? AND name=? LIMIT 1",
            (clean_site, candidate),
        ).fetchone()
        if not dup:
            return candidate
    raise HTTPException(status_code=409, detail="이름 자동생성 범위를 초과했습니다. 다른 제목을 입력해 주세요.")


def _ensure_upload_dirs() -> None:
    PHOTO_ROOT.mkdir(parents=True, exist_ok=True)
    ARCHIVE_ROOT.mkdir(parents=True, exist_ok=True)


def _save_template_items(con, template_id: int, items: List[TemplateItemPayload], ts: str) -> None:
    con.execute("DELETE FROM inspection_template_items WHERE template_id=?", (int(template_id),))
    for idx, item in enumerate(items):
        item_key = str(item.item_key or "").strip()
        item_text = str(item.item_text or "").strip()
        if not item_key or not item_text:
            continue
        con.execute(
            """
            INSERT INTO inspection_template_items(
              template_id, item_key, item_text, category, severity, sort_order,
              requires_photo, requires_note, is_active, created_at, updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                int(template_id),
                item_key,
                item_text,
                str(item.category or "").strip(),
                int(item.severity or 1),
                int(item.sort_order if item.sort_order is not None else idx),
                1 if item.requires_photo else 0,
                1 if item.requires_note else 0,
                1 if item.is_active else 0,
                ts,
                ts,
            ),
        )


def _backup_template_form_snapshot(
    con,
    *,
    template_id: int,
    actor_login: str,
    backup_source: str,
    ts: str,
) -> int:
    row = con.execute(
        """
        SELECT tp.id, tp.site_code, tp.target_id, tp.name, tp.period, tp.is_active, tp.created_by, tp.created_at, tp.updated_at,
               t.name AS target_name
        FROM inspection_templates tp
        LEFT JOIN inspection_targets t ON t.id = tp.target_id
        WHERE tp.id=?
        LIMIT 1
        """,
        (int(template_id),),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="점검표를 찾을 수 없습니다.")

    template = dict(row)
    items = [
        dict(x)
        for x in con.execute(
            """
            SELECT id, template_id, item_key, item_text, category, severity, sort_order,
                   requires_photo, requires_note, is_active, created_at, updated_at
            FROM inspection_template_items
            WHERE template_id=?
            ORDER BY sort_order ASC, id ASC
            """,
            (int(template_id),),
        ).fetchall()
    ]
    snapshot = {"template": template, "items": items}
    snapshot_json = json.dumps(snapshot, ensure_ascii=False, separators=(",", ":"))
    checksum = hashlib.sha256(snapshot_json.encode("utf-8")).hexdigest()

    con.execute(
        """
        INSERT INTO inspection_template_backups(
          site_code, target_id, target_name, template_id, template_name, period, item_count,
          snapshot_json, checksum, backup_source, created_by, created_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            str(template.get("site_code") or ""),
            int(template.get("target_id") or 0) or None,
            str(template.get("target_name") or ""),
            int(template.get("id") or 0),
            str(template.get("name") or ""),
            str(template.get("period") or ""),
            len(items),
            snapshot_json,
            checksum,
            str(backup_source or "manual"),
            str(actor_login or ""),
            ts,
        ),
    )
    return int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])


@router.get("/inspection/bootstrap")
def inspection_bootstrap(request: Request, site_code: str = Query(default="")):
    user, _token = _require_inspection_user(request)
    scoped_site = _scope_site_code(user, site_code)
    con = _connect()
    try:
        targets = [
            dict(x)
            for x in con.execute(
                """
                SELECT id, site_code, name, location, description, is_active, updated_at
                FROM inspection_targets
                WHERE site_code=? AND is_active=1
                ORDER BY name ASC
                """,
                (scoped_site,),
            ).fetchall()
        ]
        templates = [
            dict(x)
            for x in con.execute(
                """
                SELECT t.id, t.site_code, t.target_id, t.name, t.period, t.is_active, t.updated_at,
                       COALESCE(c.cnt,0) AS item_count
                FROM inspection_templates t
                LEFT JOIN (
                  SELECT template_id, COUNT(*) AS cnt
                  FROM inspection_template_items
                  WHERE is_active=1
                  GROUP BY template_id
                ) c ON c.template_id=t.id
                WHERE t.site_code=? AND t.is_active=1
                ORDER BY t.name ASC
                """,
                (scoped_site,),
            ).fetchall()
        ]
        return {
            "ok": True,
            "site_code": scoped_site,
            "user": {
                "login_id": user.get("login_id"),
                "name": user.get("name"),
                "role": user.get("role"),
                "is_admin": bool(user.get("is_admin")),
                "is_site_admin": bool(user.get("is_site_admin")),
                "can_create_run": _is_run_creator(user),
            },
            "targets": targets,
            "templates": templates,
        }
    finally:
        con.close()


@router.get("/inspection/targets")
def inspection_targets_list(
    request: Request,
    site_code: str = Query(default=""),
    active: int = Query(default=1, ge=0, le=1),
):
    user, _token = _require_inspection_user(request)
    scoped_site = _scope_site_code(user, site_code)
    con = _connect()
    try:
        rows = con.execute(
            """
            SELECT id, site_code, name, location, description, is_active, created_by, created_at, updated_at
            FROM inspection_targets
            WHERE site_code=? AND is_active=?
            ORDER BY name ASC
            """,
            (scoped_site, int(active)),
        ).fetchall()
        return {"ok": True, "items": [dict(x) for x in rows]}
    finally:
        con.close()


@router.post("/inspection/targets")
def inspection_targets_create(request: Request, payload: TargetPayload = Body(...)):
    user, _token = _require_manager_user(request)
    scoped_site = _scope_site_code(user, payload.site_code)
    ts = _now_ts()
    clean_name = str(payload.name).strip()
    clean_location = str(payload.location or "").strip()
    clean_description = str(payload.description or "").strip()
    con = _connect()
    try:
        if bool(payload.force_new):
            clean_name = _next_unique_master_name(con, "inspection_targets", scoped_site, clean_name)

        existing = con.execute(
            """
            SELECT id
            FROM inspection_targets
            WHERE site_code=? AND name=?
            LIMIT 1
            """,
            (scoped_site, clean_name),
        ).fetchone()
        if existing and not bool(payload.force_new):
            rid = int(existing["id"])
            con.execute(
                """
                UPDATE inspection_targets
                SET location=?, description=?, is_active=?, updated_at=?
                WHERE id=?
                """,
                (clean_location, clean_description, 1 if payload.is_active else 0, ts, rid),
            )
            con.commit()
            row = con.execute("SELECT * FROM inspection_targets WHERE id=?", (rid,)).fetchone()
            return {"ok": True, "item": dict(row) if row else {"id": rid}, "upserted": True}

        con.execute(
            """
            INSERT INTO inspection_targets(site_code, name, location, description, is_active, created_by, created_at, updated_at)
            VALUES(?,?,?,?,?,?,?,?)
            """,
            (
                scoped_site,
                clean_name,
                clean_location,
                clean_description,
                1 if payload.is_active else 0,
                str(user.get("login_id") or ""),
                ts,
                ts,
            ),
        )
        rid = int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        con.commit()
        row = con.execute("SELECT * FROM inspection_targets WHERE id=?", (rid,)).fetchone()
        return {"ok": True, "item": dict(row) if row else {"id": rid}}
    except Exception as e:
        raise HTTPException(status_code=409, detail=f"점검대상 생성 실패: {e}") from e
    finally:
        con.close()


@router.patch("/inspection/targets/{target_id}")
def inspection_targets_update(target_id: int, request: Request, payload: TargetUpdatePayload = Body(...)):
    user, _token = _require_manager_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        row = con.execute("SELECT * FROM inspection_targets WHERE id=? LIMIT 1", (int(target_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="점검대상을 찾을 수 없습니다.")
        item = dict(row)
        scoped_site = _scope_site_code(user, item.get("site_code"))

        new_name = str(payload.name).strip() if payload.name is not None else str(item.get("name") or "").strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="점검대상 이름은 비울 수 없습니다.")

        dup = con.execute(
            """
            SELECT id FROM inspection_targets
            WHERE site_code=? AND name=? AND id<>?
            LIMIT 1
            """,
            (scoped_site, new_name, int(target_id)),
        ).fetchone()
        if dup:
            raise HTTPException(status_code=409, detail="같은 이름의 점검대상이 이미 있습니다.")

        new_location = str(payload.location) if payload.location is not None else str(item.get("location") or "")
        new_description = str(payload.description) if payload.description is not None else str(item.get("description") or "")
        new_active = int(payload.is_active) if payload.is_active is not None else int(item.get("is_active") or 0)
        new_active = 1 if new_active == 1 else 0

        con.execute(
            """
            UPDATE inspection_targets
            SET name=?, location=?, description=?, is_active=?, updated_at=?
            WHERE id=?
            """,
            (new_name, new_location.strip(), new_description.strip(), new_active, ts, int(target_id)),
        )
        con.commit()
        out = con.execute("SELECT * FROM inspection_targets WHERE id=? LIMIT 1", (int(target_id),)).fetchone()
        return {"ok": True, "item": dict(out) if out else {"id": int(target_id)}}
    finally:
        con.close()


@router.delete("/inspection/targets/{target_id}")
def inspection_targets_delete(target_id: int, request: Request):
    user, _token = _require_manager_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        row = con.execute("SELECT * FROM inspection_targets WHERE id=? LIMIT 1", (int(target_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="점검대상을 찾을 수 없습니다.")
        item = dict(row)
        _scope_site_code(user, item.get("site_code"))
        con.execute("UPDATE inspection_targets SET is_active=0, updated_at=? WHERE id=?", (ts, int(target_id)))
        con.commit()
        return {"ok": True, "deleted": True, "target_id": int(target_id)}
    finally:
        con.close()


@router.get("/inspection/templates")
def inspection_templates_list(
    request: Request,
    site_code: str = Query(default=""),
    active: int = Query(default=1, ge=0, le=1),
    include_items: int = Query(default=1, ge=0, le=1),
):
    user, _token = _require_inspection_user(request)
    scoped_site = _scope_site_code(user, site_code)
    con = _connect()
    try:
        rows = [
            dict(x)
            for x in con.execute(
                """
                SELECT tp.*, t.name AS target_name
                FROM inspection_templates tp
                LEFT JOIN inspection_targets t ON t.id = tp.target_id
                WHERE tp.site_code=? AND tp.is_active=?
                ORDER BY tp.name ASC
                """,
                (scoped_site, int(active)),
            ).fetchall()
        ]
        if include_items:
            for row in rows:
                row["items"] = [
                    dict(it)
                    for it in con.execute(
                        """
                        SELECT *
                        FROM inspection_template_items
                        WHERE template_id=? AND is_active=1
                        ORDER BY sort_order ASC, id ASC
                        """,
                        (int(row["id"]),),
                    ).fetchall()
                ]
        return {"ok": True, "items": rows}
    finally:
        con.close()


@router.post("/inspection/templates")
def inspection_templates_create(request: Request, payload: TemplatePayload = Body(...)):
    user, _token = _require_manager_user(request)
    scoped_site = _scope_site_code(user, payload.site_code)
    ts = _now_ts()
    clean_name = str(payload.name).strip()
    clean_period = _safe_period(payload.period)
    do_backup = bool(payload.auto_backup) or bool(payload.force_new)
    con = _connect()
    try:
        target = con.execute(
            "SELECT id, site_code FROM inspection_targets WHERE id=? AND is_active=1 LIMIT 1",
            (int(payload.target_id),),
        ).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="점검대상을 찾을 수 없습니다.")
        _scope_site_code(user, target["site_code"])

        if bool(payload.force_new):
            clean_name = _next_unique_master_name(con, "inspection_templates", scoped_site, clean_name)

        existing = con.execute(
            """
            SELECT id
            FROM inspection_templates
            WHERE site_code=? AND name=?
            LIMIT 1
            """,
            (scoped_site, clean_name),
        ).fetchone()
        if existing and not bool(payload.force_new):
            template_id = int(existing["id"])
            con.execute(
                """
                UPDATE inspection_templates
                SET target_id=?, period=?, is_active=?, updated_at=?
                WHERE id=?
                """,
                (int(payload.target_id), clean_period, 1 if payload.is_active else 0, ts, template_id),
            )
            _save_template_items(con, template_id, payload.items, ts)
            backup_id = None
            if do_backup:
                backup_id = _backup_template_form_snapshot(
                    con,
                    template_id=template_id,
                    actor_login=str(user.get("login_id") or ""),
                    backup_source="auto_form" if bool(payload.force_new) else "manual_form",
                    ts=ts,
                )
            con.commit()
            return {"ok": True, "template_id": template_id, "upserted": True, "backup_id": backup_id}

        con.execute(
            """
            INSERT INTO inspection_templates(site_code, target_id, name, period, is_active, created_by, created_at, updated_at)
            VALUES(?,?,?,?,?,?,?,?)
            """,
            (
                scoped_site,
                int(payload.target_id),
                clean_name,
                clean_period,
                1 if payload.is_active else 0,
                str(user.get("login_id") or ""),
                ts,
                ts,
            ),
        )
        template_id = int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        _save_template_items(con, template_id, payload.items, ts)
        backup_id = None
        if do_backup:
            backup_id = _backup_template_form_snapshot(
                con,
                template_id=template_id,
                actor_login=str(user.get("login_id") or ""),
                backup_source="auto_form" if bool(payload.force_new) else "manual_form",
                ts=ts,
            )
        con.commit()
        return {"ok": True, "template_id": template_id, "backup_id": backup_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=409, detail=f"점검표 생성 실패: {e}") from e
    finally:
        con.close()


@router.patch("/inspection/templates/{template_id}")
def inspection_templates_update(template_id: int, request: Request, payload: TemplateUpdatePayload = Body(...)):
    user, _token = _require_manager_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        row = con.execute("SELECT * FROM inspection_templates WHERE id=? LIMIT 1", (int(template_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="점검표를 찾을 수 없습니다.")
        cur = dict(row)
        scoped_site = _scope_site_code(user, cur.get("site_code"))

        target_id = int(cur.get("target_id") or 0)
        if payload.target_id is not None:
            t = con.execute(
                "SELECT id, site_code FROM inspection_targets WHERE id=? AND is_active=1 LIMIT 1",
                (int(payload.target_id),),
            ).fetchone()
            if not t:
                raise HTTPException(status_code=404, detail="점검대상을 찾을 수 없습니다.")
            if _clean_site_code(t["site_code"]) != scoped_site:
                raise HTTPException(status_code=403, detail="같은 단지의 점검대상으로만 변경할 수 있습니다.")
            target_id = int(payload.target_id)

        new_name = str(payload.name).strip() if payload.name is not None else str(cur.get("name") or "").strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="점검표 이름은 비울 수 없습니다.")
        dup = con.execute(
            """
            SELECT id FROM inspection_templates
            WHERE site_code=? AND name=? AND id<>?
            LIMIT 1
            """,
            (scoped_site, new_name, int(template_id)),
        ).fetchone()
        if dup:
            raise HTTPException(status_code=409, detail="같은 이름의 점검표가 이미 있습니다.")

        new_period = _safe_period(payload.period) if payload.period is not None else str(cur.get("period") or "MONTHLY")
        new_active = int(payload.is_active) if payload.is_active is not None else int(cur.get("is_active") or 0)
        new_active = 1 if new_active == 1 else 0

        con.execute(
            """
            UPDATE inspection_templates
            SET target_id=?, name=?, period=?, is_active=?, updated_at=?
            WHERE id=?
            """,
            (target_id, new_name, new_period, new_active, ts, int(template_id)),
        )

        if payload.items is not None:
            _save_template_items(con, int(template_id), payload.items, ts)

        backup_id = None
        if bool(payload.auto_backup):
            backup_id = _backup_template_form_snapshot(
                con,
                template_id=int(template_id),
                actor_login=str(user.get("login_id") or ""),
                backup_source="manual_form",
                ts=ts,
            )
        con.commit()
        out = con.execute("SELECT * FROM inspection_templates WHERE id=? LIMIT 1", (int(template_id),)).fetchone()
        return {"ok": True, "item": dict(out) if out else {"id": int(template_id)}, "backup_id": backup_id}
    finally:
        con.close()


@router.delete("/inspection/templates/{template_id}")
def inspection_templates_delete(template_id: int, request: Request):
    user, _token = _require_manager_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        row = con.execute("SELECT * FROM inspection_templates WHERE id=? LIMIT 1", (int(template_id),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="점검표를 찾을 수 없습니다.")
        cur = dict(row)
        _scope_site_code(user, cur.get("site_code"))

        con.execute("UPDATE inspection_templates SET is_active=0, updated_at=? WHERE id=?", (ts, int(template_id)))
        con.execute("UPDATE inspection_template_items SET is_active=0, updated_at=? WHERE template_id=?", (ts, int(template_id)))
        con.commit()
        return {"ok": True, "deleted": True, "template_id": int(template_id)}
    finally:
        con.close()


@router.post("/inspection/runs")
def inspection_runs_create(request: Request, payload: RunCreatePayload = Body(...)):
    user, _token = _require_run_creator_user(request)
    scoped_site = _scope_site_code(user, payload.site_code)
    run_date = _safe_date(payload.run_date)
    ts = _now_ts()
    con = _connect()
    try:
        target = con.execute(
            "SELECT * FROM inspection_targets WHERE id=? AND site_code=? AND is_active=1 LIMIT 1",
            (int(payload.target_id), scoped_site),
        ).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="점검대상을 찾을 수 없습니다.")
        template = con.execute(
            "SELECT * FROM inspection_templates WHERE id=? AND site_code=? AND is_active=1 LIMIT 1",
            (int(payload.template_id), scoped_site),
        ).fetchone()
        if not template:
            raise HTTPException(status_code=404, detail="점검표를 찾을 수 없습니다.")
        items = con.execute(
            """
            SELECT * FROM inspection_template_items
            WHERE template_id=? AND is_active=1
            ORDER BY sort_order ASC, id ASC
            """,
            (int(payload.template_id),),
        ).fetchall()
        if not items:
            raise HTTPException(status_code=409, detail="점검표에 활성 항목이 없습니다.")

        run_code = _next_run_code(con, run_date)
        con.execute(
            """
            INSERT INTO inspection_runs(
              run_code, site_code, target_id, template_id, inspector_user_id, inspector_login, inspector_name,
              run_date, status, note, created_at, updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                run_code,
                scoped_site,
                int(payload.target_id),
                int(payload.template_id),
                int(user.get("id") or 0) or None,
                str(user.get("login_id") or ""),
                str(user.get("name") or ""),
                run_date,
                "DRAFT",
                str(payload.note or "").strip(),
                ts,
                ts,
            ),
        )
        run_id = int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        for item in items:
            con.execute(
                """
                INSERT INTO inspection_run_items(
                  run_id, template_item_id, item_key, item_text, category, severity,
                  requires_photo, requires_note, result, note, updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    run_id,
                    int(item["id"]),
                    str(item["item_key"] or ""),
                    str(item["item_text"] or ""),
                    str(item["category"] or ""),
                    int(item["severity"] or 1),
                    1 if int(item["requires_photo"] or 0) == 1 else 0,
                    1 if int(item["requires_note"] or 0) == 1 else 0,
                    "NA",
                    "",
                    ts,
                ),
            )
        con.commit()
        return {"ok": True, "run_id": run_id, "run_code": run_code}
    finally:
        con.close()


@router.get("/inspection/runs")
def inspection_runs_list(
    request: Request,
    site_code: str = Query(default=""),
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
    status: str = Query(default=""),
    target_id: int = Query(default=0, ge=0),
    template_id: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        rows = _query_runs_rows(
            con=con,
            user=user,
            site_code=site_code,
            date_from=date_from,
            date_to=date_to,
            status=status,
            target_id=target_id,
            template_id=template_id,
            limit=limit,
        )
        return {"ok": True, "items": rows}
    finally:
        con.close()


def _query_runs_rows(
    *,
    con,
    user: Dict[str, Any],
    site_code: str,
    date_from: str,
    date_to: str,
    status: str,
    target_id: int,
    template_id: int,
    limit: int,
) -> List[Dict[str, Any]]:
    scoped_site = _scope_site_code(user, site_code)
    clauses = ["r.site_code=?"]
    params: List[Any] = [scoped_site]
    if date_from:
        clauses.append("r.run_date>=?")
        params.append(_safe_date(date_from))
    if date_to:
        clauses.append("r.run_date<=?")
        params.append(_safe_date(date_to))
    if status:
        clauses.append("upper(r.status)=?")
        params.append(str(status or "").strip().upper())
    if target_id > 0:
        clauses.append("r.target_id=?")
        params.append(int(target_id))
    if template_id > 0:
        clauses.append("r.template_id=?")
        params.append(int(template_id))
    if (not _is_admin(user)) and (not _is_site_admin(user)):
        clauses.append("lower(r.inspector_login)=?")
        params.append(str(user.get("login_id") or "").strip().lower())
    where_sql = " AND ".join(clauses)
    rows = con.execute(
        f"""
        SELECT r.*,
               t.name AS target_name,
               tp.name AS template_name,
               COALESCE(s.item_count, 0) AS item_count,
               COALESCE(s.non_count, 0) AS noncompliant_count
        FROM inspection_runs r
        LEFT JOIN inspection_targets t ON t.id=r.target_id
        LEFT JOIN inspection_templates tp ON tp.id=r.template_id
        LEFT JOIN (
          SELECT run_id,
                 COUNT(*) AS item_count,
                 SUM(CASE WHEN upper(result)='NONCOMPLIANT' THEN 1 ELSE 0 END) AS non_count
          FROM inspection_run_items
          GROUP BY run_id
        ) s ON s.run_id=r.id
        WHERE {where_sql}
        ORDER BY r.run_date DESC, r.id DESC
        LIMIT ?
        """,
        tuple(params + [int(limit)]),
    ).fetchall()
    return [dict(x) for x in rows]


def _run_status_label(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if raw == "DRAFT":
        return "작성중"
    if raw == "SUBMITTED":
        return "제출"
    if raw == "REJECTED":
        return "반려"
    if raw == "ARCHIVED":
        return "보관"
    return raw or "-"


@router.get("/inspection/runs/export.xlsx")
def inspection_runs_export_xlsx(
    request: Request,
    site_code: str = Query(default=""),
    date_from: str = Query(default=""),
    date_to: str = Query(default=""),
    status: str = Query(default=""),
    target_id: int = Query(default=0, ge=0),
    template_id: int = Query(default=0, ge=0),
    limit: int = Query(default=2000, ge=1, le=10000),
):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        rows = _query_runs_rows(
            con=con,
            user=user,
            site_code=site_code,
            date_from=date_from,
            date_to=date_to,
            status=status,
            target_id=target_id,
            template_id=template_id,
            limit=limit,
        )
    finally:
        con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "점검목록"
    headers = [
        "점검ID",
        "점검코드",
        "단지코드",
        "점검일자",
        "상태",
        "점검대상",
        "점검표",
        "점검자ID",
        "점검자명",
        "항목수",
        "부적합건수",
        "메모",
        "생성일",
        "갱신일",
        "완료일",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                int(r.get("id") or 0),
                str(r.get("run_code") or ""),
                str(r.get("site_code") or ""),
                str(r.get("run_date") or ""),
                _run_status_label(r.get("status")),
                str(r.get("target_name") or ""),
                str(r.get("template_name") or ""),
                str(r.get("inspector_login") or ""),
                str(r.get("inspector_name") or ""),
                int(r.get("item_count") or 0),
                int(r.get("noncompliant_count") or 0),
                str(r.get("note") or ""),
                str(r.get("created_at") or ""),
                str(r.get("updated_at") or ""),
                str(r.get("completed_at") or ""),
            ]
        )

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(42, max_len + 2))
    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    file_name = f"inspection_runs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/inspection/runs/{run_id}")
def inspection_runs_detail(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        detail = _query_run_detail_data(con=con, user=user, run_id=run_id)
        return {"ok": True, **detail}
    finally:
        con.close()


def _query_run_detail_data(*, con, user: Dict[str, Any], run_id: int) -> Dict[str, Any]:
    run_row = con.execute(
        """
        SELECT r.*, t.name AS target_name, tp.name AS template_name
        FROM inspection_runs r
        LEFT JOIN inspection_targets t ON t.id=r.target_id
        LEFT JOIN inspection_templates tp ON tp.id=r.template_id
        WHERE r.id=?
        LIMIT 1
        """,
        (int(run_id),),
    ).fetchone()
    if not run_row:
        raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
    run = dict(run_row)
    _run_scope_check(user, run)
    items = [
        dict(x)
        for x in con.execute(
            """
            SELECT id, item_key, item_text, category, severity, requires_photo, requires_note,
                   result, note, photo_path, photo_name, updated_at
            FROM inspection_run_items
            WHERE run_id=?
            ORDER BY id ASC
            """,
            (int(run_id),),
        ).fetchall()
    ]
    approvals = [
        dict(x)
        for x in con.execute(
            """
            SELECT step_no, approver_login, approver_name, decision, comment, decided_at, created_at
            FROM inspection_approvals
            WHERE run_id=?
            ORDER BY step_no ASC
            """,
            (int(run_id),),
        ).fetchall()
    ]
    archive = con.execute(
        "SELECT run_id, pdf_relpath, checksum, archived_by, archived_at FROM inspection_archives WHERE run_id=? LIMIT 1",
        (int(run_id),),
    ).fetchone()
    return {"run": run, "items": items, "approvals": approvals, "archive": dict(archive) if archive else None}


@router.get("/inspection/runs/{run_id}/export.xlsx")
def inspection_runs_detail_export_xlsx(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        detail = _query_run_detail_data(con=con, user=user, run_id=run_id)
    finally:
        con.close()

    run = detail.get("run") or {}
    items = detail.get("items") or []
    approvals = detail.get("approvals") or []
    archive = detail.get("archive") or {}

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "점검개요"
    ws_meta.append(["항목", "값"])
    ws_meta_rows = [
        ("점검ID", int(run.get("id") or 0)),
        ("점검코드", str(run.get("run_code") or "")),
        ("단지코드", str(run.get("site_code") or "")),
        ("점검일자", str(run.get("run_date") or "")),
        ("상태", _run_status_label(run.get("status"))),
        ("점검대상", str(run.get("target_name") or "")),
        ("점검표", str(run.get("template_name") or "")),
        ("점검자ID", str(run.get("inspector_login") or "")),
        ("점검자명", str(run.get("inspector_name") or "")),
        ("메모", str(run.get("note") or "")),
        ("생성일", str(run.get("created_at") or "")),
        ("갱신일", str(run.get("updated_at") or "")),
        ("완료일", str(run.get("completed_at") or "")),
    ]
    for key, value in ws_meta_rows:
        ws_meta.append([key, value])
    ws_meta.column_dimensions["A"].width = 18
    ws_meta.column_dimensions["B"].width = 52

    ws_items = wb.create_sheet("점검항목")
    item_headers = ["ID", "분류", "항목", "결과", "메모", "사진파일", "갱신일"]
    ws_items.append(item_headers)
    for row in items:
        ws_items.append(
            [
                int(row.get("id") or 0),
                str(row.get("category") or ""),
                str(row.get("item_text") or row.get("item_key") or ""),
                _result_label(row.get("result") or "NA"),
                str(row.get("note") or ""),
                str(row.get("photo_name") or ""),
                str(row.get("updated_at") or ""),
            ]
        )
    for col_idx, header in enumerate(item_headers, start=1):
        max_len = len(str(header))
        for row in ws_items.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws_items.column_dimensions[get_column_letter(col_idx)].width = max(10, min(42, max_len + 2))
    ws_items.freeze_panes = "A2"

    ws_approvals = wb.create_sheet("결재이력")
    approval_headers = ["단계", "결재자ID", "결재자명", "결정", "코멘트", "결정시각", "생성시각"]
    ws_approvals.append(approval_headers)
    for row in approvals:
        ws_approvals.append(
            [
                int(row.get("step_no") or 0),
                str(row.get("approver_login") or ""),
                str(row.get("approver_name") or ""),
                str(row.get("decision") or "PENDING"),
                str(row.get("comment") or ""),
                str(row.get("decided_at") or ""),
                str(row.get("created_at") or ""),
            ]
        )
    for col_idx, header in enumerate(approval_headers, start=1):
        max_len = len(str(header))
        for row in ws_approvals.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws_approvals.column_dimensions[get_column_letter(col_idx)].width = max(10, min(38, max_len + 2))
    ws_approvals.freeze_panes = "A2"

    if archive:
        ws_archive = wb.create_sheet("보관정보")
        ws_archive.append(["항목", "값"])
        ws_archive.append(["PDF 경로", str(archive.get("pdf_relpath") or "")])
        ws_archive.append(["체크섬", str(archive.get("checksum") or "")])
        ws_archive.append(["보관자", str(archive.get("archived_by") or "")])
        ws_archive.append(["보관시각", str(archive.get("archived_at") or "")])
        ws_archive.column_dimensions["A"].width = 18
        ws_archive.column_dimensions["B"].width = 60

    bio = BytesIO()
    wb.save(bio)
    run_code = str(run.get("run_code") or f"run_{int(run_id)}")
    safe_run_code = re.sub(r"[^A-Za-z0-9._-]+", "_", run_code)
    file_name = f"inspection_run_{safe_run_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.patch("/inspection/runs/{run_id}/items")
def inspection_runs_items_patch(run_id: int, request: Request, payload: RunItemsPatchRequest = Body(...)):
    user, _token = _require_inspection_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (int(run_id),)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        if str(run.get("status") or "").upper() not in {"DRAFT", "REJECTED"}:
            raise HTTPException(status_code=409, detail="작성중/반려 상태에서만 수정할 수 있습니다.")
        if not _can_edit_run(user, run):
            raise HTTPException(status_code=403, detail="본인 점검만 수정할 수 있습니다.")
        for item in payload.items or []:
            row = con.execute(
                "SELECT id FROM inspection_run_items WHERE id=? AND run_id=? LIMIT 1",
                (int(item.id), int(run_id)),
            ).fetchone()
            if not row:
                continue
            con.execute(
                """
                UPDATE inspection_run_items
                SET result=?, note=?, updated_at=?
                WHERE id=? AND run_id=?
                """,
                (
                    _safe_result(item.result),
                    str(item.note or "").strip(),
                    ts,
                    int(item.id),
                    int(run_id),
                ),
            )
        con.execute("UPDATE inspection_runs SET updated_at=? WHERE id=?", (ts, int(run_id)))
        con.commit()
        return {"ok": True}
    finally:
        con.close()


@router.post("/inspection/runs/{run_id}/items/{item_id}/photo")
async def inspection_runs_item_photo_upload(
    run_id: int,
    item_id: int,
    request: Request,
    photo: UploadFile = File(...),
):
    user, _token = _require_inspection_user(request)
    run_id = int(run_id)
    item_id = int(item_id)
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (run_id,)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        if str(run.get("status") or "").upper() not in {"DRAFT", "REJECTED"}:
            raise HTTPException(status_code=409, detail="작성중/반려 상태에서만 사진을 업로드할 수 있습니다.")
        if not _can_edit_run(user, run):
            raise HTTPException(status_code=403, detail="본인 점검만 수정할 수 있습니다.")

        item_row = con.execute(
            "SELECT id FROM inspection_run_items WHERE id=? AND run_id=? LIMIT 1",
            (item_id, run_id),
        ).fetchone()
        if not item_row:
            raise HTTPException(status_code=404, detail="점검 항목을 찾을 수 없습니다.")

        file_name = str(photo.filename or "photo.jpg")
        raw = await photo.read()
        if not raw:
            raise HTTPException(status_code=400, detail="비어 있는 파일은 업로드할 수 없습니다.")
        if len(raw) > PHOTO_MAX_FILE_BYTES:
            raise HTTPException(status_code=413, detail=f"사진 용량은 {PHOTO_MAX_FILE_BYTES} 바이트 이하여야 합니다.")

        _ensure_upload_dirs()
        ext = Path(file_name).suffix.lower() or ".jpg"
        run_code = str(run.get("run_code") or f"run-{run_id}")
        safe_run_code = re.sub(r"[^A-Za-z0-9_-]+", "_", run_code)
        save_dir = (PHOTO_ROOT / safe_run_code).resolve()
        save_dir.mkdir(parents=True, exist_ok=True)
        save_name = f"{uuid.uuid4().hex}{ext}"
        abs_path = (save_dir / save_name).resolve()
        rel_path = f"inspection/photos/{safe_run_code}/{save_name}"
        abs_path.write_bytes(raw)

        ts = _now_ts()
        con.execute(
            """
            UPDATE inspection_run_items
            SET photo_path=?, photo_name=?, updated_at=?
            WHERE id=? AND run_id=?
            """,
            (rel_path, file_name, ts, item_id, run_id),
        )
        con.execute("UPDATE inspection_runs SET updated_at=? WHERE id=?", (ts, run_id))
        con.commit()
        return {"ok": True, "photo_path": rel_path, "photo_name": file_name}
    finally:
        con.close()


@router.get("/inspection/runs/{run_id}/items/{item_id}/photo")
def inspection_runs_item_photo_get(run_id: int, item_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (int(run_id),)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        item = con.execute(
            "SELECT photo_path, photo_name FROM inspection_run_items WHERE id=? AND run_id=? LIMIT 1",
            (int(item_id), int(run_id)),
        ).fetchone()
        if not item or not item["photo_path"]:
            raise HTTPException(status_code=404, detail="사진이 없습니다.")
        rel = str(item["photo_path"] or "").strip()
        abs_path = (ROOT_DIR / "uploads" / rel).resolve()
        if not abs_path.exists() or not abs_path.is_file():
            raise HTTPException(status_code=404, detail="사진 파일을 찾을 수 없습니다.")
        return FileResponse(path=str(abs_path), filename=str(item["photo_name"] or abs_path.name))
    finally:
        con.close()


def _resolve_approval_chain(con, run: Dict[str, Any]) -> List[Dict[str, Any]]:
    site_code = _clean_site_code(run.get("site_code"))
    inspector_login = str(run.get("inspector_login") or "").strip().lower()
    candidates: List[Dict[str, Any]] = []

    site_admin = con.execute(
        """
        SELECT id, login_id, name
        FROM staff_users
        WHERE is_active=1 AND is_site_admin=1 AND site_code=?
        ORDER BY id ASC
        LIMIT 1
        """,
        (site_code,),
    ).fetchone()
    if site_admin:
        candidates.append(dict(site_admin))

    admins = con.execute(
        """
        SELECT id, login_id, name
        FROM staff_users
        WHERE is_active=1 AND is_admin=1
        ORDER BY CASE WHEN lower(COALESCE(admin_scope,''))='super_admin' THEN 0 ELSE 1 END, id ASC
        """,
    ).fetchall()
    for row in admins:
        candidates.append(dict(row))

    out: List[Dict[str, Any]] = []
    used = set()
    for c in candidates:
        login = str(c.get("login_id") or "").strip().lower()
        if not login or login in used or login == inspector_login:
            continue
        used.add(login)
        out.append(c)
        if len(out) >= 2:
            break
    return out


def _run_snapshot(con, run_id: int) -> Dict[str, Any]:
    row = con.execute(
        """
        SELECT r.*, t.name AS target_name, tp.name AS template_name
        FROM inspection_runs r
        LEFT JOIN inspection_targets t ON t.id = r.target_id
        LEFT JOIN inspection_templates tp ON tp.id = r.template_id
        WHERE r.id=?
        LIMIT 1
        """,
        (int(run_id),),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
    run = dict(row)
    items = [dict(x) for x in con.execute("SELECT * FROM inspection_run_items WHERE run_id=? ORDER BY id ASC", (int(run_id),)).fetchall()]
    approvals = [dict(x) for x in con.execute("SELECT * FROM inspection_approvals WHERE run_id=? ORDER BY step_no ASC", (int(run_id),)).fetchall()]
    return {"run": run, "items": items, "approvals": approvals}


def _write_archive_pdf(snapshot: Dict[str, Any], run_code: str) -> str:
    _ensure_upload_dirs()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{run_code}_{stamp}.pdf"
    abs_path = (ARCHIVE_ROOT / file_name).resolve()

    c = canvas.Canvas(str(abs_path), pagesize=A4)
    width, height = A4
    y = height - 40

    def line(txt: str, step: int = 14):
        nonlocal y
        if y < 50:
            c.showPage()
            y = height - 40
        c.drawString(32, y, str(txt))
        y -= step

    run = snapshot.get("run") or {}
    line("안전점검 보고서")
    line(f"점검코드: {run.get('run_code') or '-'}")
    line(f"단지코드: {run.get('site_code') or '-'}")
    line(f"점검일자: {run.get('run_date') or '-'}")
    line(f"점검자: {run.get('inspector_name') or run.get('inspector_login') or '-'}")
    line(f"상태: {run.get('status') or '-'}")
    line(f"점검대상: {run.get('target_name') or '-'}")
    line(f"템플릿: {run.get('template_name') or '-'}")
    line("")
    line("[점검 항목]")
    for idx, item in enumerate(snapshot.get("items") or [], start=1):
        label = str(item.get("item_text") or item.get("item_key") or f"항목{idx}")
        result = _result_label(item.get("result") or "NA")
        note = str(item.get("note") or "").strip()
        line(f"{idx}. {label} / 결과: {result}")
        if note:
            line(f"   메모: {note}", 12)
    line("")
    line("[결재 이력]")
    for ap in snapshot.get("approvals") or []:
        line(
            f"step {ap.get('step_no')} / {ap.get('approver_name') or ap.get('approver_login') or '-'} / "
            f"{ap.get('decision') or 'PENDING'} / {ap.get('decided_at') or '-'}"
        )
    c.save()
    return f"inspection/archives/{file_name}"


def _archive_run(con, run_id: int, archived_by: str) -> Dict[str, Any]:
    ts = _now_ts()
    snapshot = _run_snapshot(con, run_id)
    run = snapshot["run"]
    run_code = str(run.get("run_code") or f"INSP-{run_id}")
    snapshot_json = json.dumps(snapshot, ensure_ascii=False, separators=(",", ":"))
    checksum = hashlib.sha256(snapshot_json.encode("utf-8")).hexdigest()
    pdf_relpath = _write_archive_pdf(snapshot, run_code)

    existing = con.execute("SELECT id FROM inspection_archives WHERE run_id=? LIMIT 1", (int(run_id),)).fetchone()
    if existing:
        con.execute(
            """
            UPDATE inspection_archives
            SET pdf_relpath=?, snapshot_json=?, checksum=?, archived_by=?, archived_at=?, updated_at=?
            WHERE run_id=?
            """,
            (pdf_relpath, snapshot_json, checksum, str(archived_by or ""), ts, ts, int(run_id)),
        )
    else:
        con.execute(
            """
            INSERT INTO inspection_archives(run_id, site_code, pdf_relpath, snapshot_json, checksum, archived_by, archived_at, created_at, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (int(run_id), str(run.get("site_code") or ""), pdf_relpath, snapshot_json, checksum, str(archived_by or ""), ts, ts, ts),
        )
    con.execute(
        "UPDATE inspection_runs SET status='ARCHIVED', completed_at=COALESCE(completed_at, ?), approval_step=NULL, updated_at=? WHERE id=?",
        (ts, ts, int(run_id)),
    )
    return {"pdf_relpath": pdf_relpath, "checksum": checksum, "archived_at": ts}


@router.post("/inspection/runs/{run_id}/submit")
def inspection_runs_submit(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    ts = _now_ts()
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (int(run_id),)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        if str(run.get("status") or "").upper() not in {"DRAFT", "REJECTED"}:
            raise HTTPException(status_code=409, detail="작성중/반려 상태에서만 제출할 수 있습니다.")
        if not _can_edit_run(user, run):
            raise HTTPException(status_code=403, detail="본인 점검만 제출할 수 있습니다.")

        rows = con.execute(
            """
            SELECT item_text, requires_photo, requires_note, result, note, photo_path
            FROM inspection_run_items
            WHERE run_id=?
            ORDER BY id ASC
            """,
            (int(run_id),),
        ).fetchall()
        for row in rows:
            if _safe_result(row["result"]) != "NONCOMPLIANT":
                continue
            note = str(row["note"] or "").strip()
            photo_path = str(row["photo_path"] or "").strip()
            if int(row["requires_note"] or 0) == 1 and not note:
                raise HTTPException(status_code=409, detail=f"부적합 항목 메모가 필요합니다: {row['item_text']}")
            if int(row["requires_photo"] or 0) == 1 and not photo_path:
                raise HTTPException(status_code=409, detail=f"부적합 항목 사진이 필요합니다: {row['item_text']}")

        approvers = _resolve_approval_chain(con, run)
        if not approvers:
            raise HTTPException(status_code=409, detail="결재자를 찾을 수 없습니다. 관리자 계정을 확인하세요.")
        con.execute("DELETE FROM inspection_approvals WHERE run_id=?", (int(run_id),))
        step_no = 1
        for ap in approvers:
            con.execute(
                """
                INSERT INTO inspection_approvals(
                  run_id, step_no, approver_user_id, approver_login, approver_name, decision, comment, decided_at, created_at, updated_at
                ) VALUES(?,?,?,?,?,'PENDING','',NULL,?,?)
                """,
                (int(run_id), step_no, int(ap.get("id") or 0) or None, str(ap.get("login_id") or ""), str(ap.get("name") or ""), ts, ts),
            )
            step_no += 1
        con.execute("UPDATE inspection_runs SET status='SUBMITTED', submitted_at=?, approval_step=1, updated_at=? WHERE id=?", (ts, ts, int(run_id)))
        con.commit()
        return {"ok": True, "status": "SUBMITTED", "approval_steps": len(approvers)}
    finally:
        con.close()


def _decide_run(run_id: int, user: Dict[str, Any], comment: str, decision: str, approval_code: str) -> Dict[str, Any]:
    decision = str(decision or "").strip().upper()
    if decision not in {"APPROVED", "REJECTED"}:
        raise HTTPException(status_code=400, detail="invalid decision")
    ts = _now_ts()
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (int(run_id),)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        if str(run.get("status") or "").upper() != "SUBMITTED":
            raise HTTPException(status_code=409, detail="제출 상태에서만 결재할 수 있습니다.")
        pending = con.execute(
            "SELECT * FROM inspection_approvals WHERE run_id=? AND decision='PENDING' ORDER BY step_no ASC LIMIT 1",
            (int(run_id),),
        ).fetchone()
        if not pending:
            raise HTTPException(status_code=409, detail="대기 중인 결재 단계가 없습니다.")
        p = dict(pending)
        actor_login = str(user.get("login_id") or "").strip().lower()
        if (not _is_admin(user)) and actor_login != str(p.get("approver_login") or "").strip().lower():
            raise HTTPException(status_code=403, detail="현재 단계 결재 권한이 없습니다.")

        # required code check (issued by /approval-code endpoint)
        expected_hash = str(p.get("approval_code_hash") or "").strip()
        expected_actor = str(p.get("approval_code_actor_login") or "").strip().lower()
        expected_expires = str(p.get("approval_code_expires_at") or "").strip()
        used_at = str(p.get("approval_code_used_at") or "").strip()
        if not expected_hash or not expected_expires:
            raise HTTPException(status_code=409, detail="승인코드를 먼저 발급해 주세요.")
        if used_at:
            raise HTTPException(status_code=409, detail="이미 사용된 승인코드입니다. 다시 발급해 주세요.")
        if expected_expires < ts:
            raise HTTPException(status_code=409, detail="승인코드가 만료되었습니다. 다시 발급해 주세요.")
        if expected_actor and expected_actor != actor_login:
            raise HTTPException(status_code=403, detail="해당 승인코드는 다른 결재자에게 발급되었습니다.")

        input_code = _normalize_approval_code(approval_code)
        if not input_code:
            raise HTTPException(status_code=400, detail="승인코드를 입력해 주세요.")
        actual_hash = _approval_code_hash(
            run_id=int(run_id),
            step_no=int(p.get("step_no") or 0),
            actor_login=actor_login,
            code=input_code,
        )
        if actual_hash != expected_hash:
            attempts = int(p.get("approval_code_attempts") or 0) + 1
            con.execute(
                """
                UPDATE inspection_approvals
                SET approval_code_attempts=?, updated_at=?,
                    approval_code_hash=CASE WHEN ?>=5 THEN NULL ELSE approval_code_hash END,
                    approval_code_actor_login=CASE WHEN ?>=5 THEN NULL ELSE approval_code_actor_login END,
                    approval_code_expires_at=CASE WHEN ?>=5 THEN NULL ELSE approval_code_expires_at END
                WHERE id=?
                """,
                (attempts, ts, attempts, attempts, attempts, int(p["id"])),
            )
            con.commit()
            if attempts >= 5:
                raise HTTPException(status_code=409, detail="승인코드 입력 오류가 누적되었습니다. 코드를 다시 발급해 주세요.")
            raise HTTPException(status_code=400, detail="승인코드가 일치하지 않습니다.")

        con.execute(
            """
            UPDATE inspection_approvals
            SET decision=?, comment=?, decided_at=?, updated_at=?,
                approval_code_used_at=?, approval_code_attempts=0
            WHERE id=?
            """,
            (decision, str(comment or "").strip(), ts, ts, ts, int(p["id"])),
        )
        if decision == "REJECTED":
            con.execute("UPDATE inspection_runs SET status='REJECTED', approval_step=NULL, updated_at=? WHERE id=?", (ts, ts, int(run_id)))
            con.commit()
            return {"ok": True, "status": "REJECTED"}

        next_pending = con.execute(
            "SELECT step_no FROM inspection_approvals WHERE run_id=? AND decision='PENDING' ORDER BY step_no ASC LIMIT 1",
            (int(run_id),),
        ).fetchone()
        if next_pending:
            con.execute("UPDATE inspection_runs SET approval_step=?, updated_at=? WHERE id=?", (int(next_pending["step_no"]), ts, int(run_id)))
            con.commit()
            return {"ok": True, "status": "SUBMITTED", "approval_step": int(next_pending["step_no"])}

        con.execute("UPDATE inspection_runs SET status='APPROVED', completed_at=?, approval_step=NULL, updated_at=? WHERE id=?", (ts, ts, int(run_id)))
        archive = _archive_run(con, int(run_id), str(user.get("login_id") or ""))
        con.commit()
        return {"ok": True, "status": "ARCHIVED", "archive": archive}
    finally:
        con.close()


@router.post("/inspection/runs/{run_id}/approve")
def inspection_runs_approve(run_id: int, request: Request, payload: DecisionPayload = Body(default={"comment": ""})):
    user, _token = _require_inspection_user(request)
    return _decide_run(int(run_id), user, str(payload.comment or ""), "APPROVED", str(payload.approval_code or ""))


@router.post("/inspection/runs/{run_id}/reject")
def inspection_runs_reject(run_id: int, request: Request, payload: DecisionPayload = Body(default={"comment": ""})):
    user, _token = _require_inspection_user(request)
    return _decide_run(int(run_id), user, str(payload.comment or ""), "REJECTED", str(payload.approval_code or ""))


@router.post("/inspection/runs/{run_id}/approval-code")
def inspection_runs_issue_approval_code(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    ts = _now_ts()
    actor_login = str(user.get("login_id") or "").strip().lower()
    con = _connect()
    try:
        run_row = con.execute("SELECT * FROM inspection_runs WHERE id=? LIMIT 1", (int(run_id),)).fetchone()
        if not run_row:
            raise HTTPException(status_code=404, detail="점검 실행내역을 찾을 수 없습니다.")
        run = dict(run_row)
        _run_scope_check(user, run)
        if str(run.get("status") or "").upper() != "SUBMITTED":
            raise HTTPException(status_code=409, detail="제출 상태에서만 승인코드를 발급할 수 있습니다.")

        pending = con.execute(
            "SELECT * FROM inspection_approvals WHERE run_id=? AND decision='PENDING' ORDER BY step_no ASC LIMIT 1",
            (int(run_id),),
        ).fetchone()
        if not pending:
            raise HTTPException(status_code=409, detail="대기 중인 결재 단계가 없습니다.")
        p = dict(pending)
        if (not _is_admin(user)) and actor_login != str(p.get("approver_login") or "").strip().lower():
            raise HTTPException(status_code=403, detail="현재 단계 결재 권한이 없습니다.")

        code = _issue_random_approval_code()
        expires_at = (datetime.now() + timedelta(minutes=APPROVAL_CODE_TTL_MINUTES)).replace(microsecond=0).isoformat(sep=" ")
        code_hash = _approval_code_hash(
            run_id=int(run_id),
            step_no=int(p.get("step_no") or 0),
            actor_login=actor_login,
            code=code,
        )
        con.execute(
            """
            UPDATE inspection_approvals
            SET approval_code_hash=?, approval_code_actor_login=?, approval_code_expires_at=?,
                approval_code_used_at=NULL, approval_code_attempts=0, approval_code_last_issued_at=?, updated_at=?
            WHERE id=?
            """,
            (code_hash, actor_login, expires_at, ts, ts, int(p["id"])),
        )
        con.commit()
        return {
            "ok": True,
            "run_id": int(run_id),
            "step_no": int(p.get("step_no") or 0),
            "approval_code": code,
            "expires_at": expires_at,
            "ttl_minutes": int(APPROVAL_CODE_TTL_MINUTES),
        }
    finally:
        con.close()


@router.get("/inspection/archives")
def inspection_archives_list(request: Request, site_code: str = Query(default=""), date_from: str = Query(default=""), date_to: str = Query(default=""), limit: int = Query(default=100, ge=1, le=500)):
    user, _token = _require_inspection_user(request)
    scoped_site = _scope_site_code(user, site_code)
    clauses = ["a.site_code=?"]
    params: List[Any] = [scoped_site]
    if date_from:
        clauses.append("substr(a.archived_at,1,10)>=?")
        params.append(_safe_date(date_from))
    if date_to:
        clauses.append("substr(a.archived_at,1,10)<=?")
        params.append(_safe_date(date_to))
    if (not _is_admin(user)) and (not _is_site_admin(user)):
        clauses.append("lower(r.inspector_login)=?")
        params.append(str(user.get("login_id") or "").strip().lower())
    where_sql = " AND ".join(clauses)
    con = _connect()
    try:
        rows = con.execute(
            f"""
            SELECT a.run_id, a.site_code, a.pdf_relpath, a.checksum, a.archived_by, a.archived_at,
                   r.run_code, r.run_date, r.inspector_login, r.inspector_name,
                   t.name AS target_name, tp.name AS template_name
            FROM inspection_archives a
            JOIN inspection_runs r ON r.id = a.run_id
            LEFT JOIN inspection_targets t ON t.id = r.target_id
            LEFT JOIN inspection_templates tp ON tp.id = r.template_id
            WHERE {where_sql}
            ORDER BY a.archived_at DESC, a.id DESC
            LIMIT ?
            """,
            tuple(params + [int(limit)]),
        ).fetchall()
        return {"ok": True, "items": [dict(x) for x in rows]}
    finally:
        con.close()


@router.get("/inspection/archives/{run_id}/verify")
def inspection_archives_verify(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        row = con.execute(
            """
            SELECT a.*, r.site_code, r.inspector_login
            FROM inspection_archives a
            JOIN inspection_runs r ON r.id = a.run_id
            WHERE a.run_id=?
            LIMIT 1
            """,
            (int(run_id),),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="보관본을 찾을 수 없습니다.")
        item = dict(row)
        _scope_site_code(user, item.get("site_code"))
        if (not _is_admin(user)) and (not _is_site_admin(user)):
            if str(item.get("inspector_login") or "").strip().lower() != str(user.get("login_id") or "").strip().lower():
                raise HTTPException(status_code=403, detail="본인 보관본만 조회할 수 있습니다.")
        snapshot_json = str(item.get("snapshot_json") or "")
        computed = hashlib.sha256(snapshot_json.encode("utf-8")).hexdigest()
        return {"ok": True, "run_id": int(run_id), "stored_checksum": str(item.get("checksum") or ""), "computed_checksum": computed, "valid": computed == str(item.get("checksum") or "")}
    finally:
        con.close()


@router.get("/inspection/archives/{run_id}/pdf")
def inspection_archives_pdf(run_id: int, request: Request):
    user, _token = _require_inspection_user(request)
    con = _connect()
    try:
        row = con.execute(
            """
            SELECT a.pdf_relpath, a.site_code, r.run_code, r.inspector_login
            FROM inspection_archives a
            JOIN inspection_runs r ON r.id = a.run_id
            WHERE a.run_id=?
            LIMIT 1
            """,
            (int(run_id),),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="보관 PDF를 찾을 수 없습니다.")
        item = dict(row)
        _scope_site_code(user, item.get("site_code"))
        if (not _is_admin(user)) and (not _is_site_admin(user)):
            if str(item.get("inspector_login") or "").strip().lower() != str(user.get("login_id") or "").strip().lower():
                raise HTTPException(status_code=403, detail="본인 보관본만 조회할 수 있습니다.")
        rel = str(item.get("pdf_relpath") or "").strip()
        if not rel:
            raise HTTPException(status_code=404, detail="보관 PDF가 없습니다.")
        abs_path = (ROOT_DIR / "uploads" / rel).resolve()
        if not abs_path.exists() or not abs_path.is_file():
            raise HTTPException(status_code=404, detail="보관 PDF 파일을 찾을 수 없습니다.")
        run_code = str(item.get("run_code") or f"inspection_{int(run_id)}")
        return FileResponse(path=str(abs_path), filename=f"{run_code}.pdf", media_type="application/pdf")
    finally:
        con.close()


@router.get("/inspection/stats")
def inspection_stats(request: Request, site_code: str = Query(default=""), date_from: str = Query(default=""), date_to: str = Query(default="")):
    user, _token = _require_inspection_user(request)
    scoped_site = _scope_site_code(user, site_code)
    clauses = ["r.site_code=?"]
    params: List[Any] = [scoped_site]
    if date_from:
        clauses.append("r.run_date>=?")
        params.append(_safe_date(date_from))
    if date_to:
        clauses.append("r.run_date<=?")
        params.append(_safe_date(date_to))
    if (not _is_admin(user)) and (not _is_site_admin(user)):
        clauses.append("lower(r.inspector_login)=?")
        params.append(str(user.get("login_id") or "").strip().lower())
    where_sql = " AND ".join(clauses)
    con = _connect()
    try:
        totals = con.execute(
            f"""
            SELECT
              COUNT(*) AS total_runs,
              SUM(CASE WHEN upper(r.status)='ARCHIVED' THEN 1 ELSE 0 END) AS archived_runs,
              SUM(CASE WHEN upper(r.status)='SUBMITTED' THEN 1 ELSE 0 END) AS submitted_runs,
              SUM(CASE WHEN upper(r.status)='REJECTED' THEN 1 ELSE 0 END) AS rejected_runs
            FROM inspection_runs r
            WHERE {where_sql}
            """,
            tuple(params),
        ).fetchone()
        top_rows = con.execute(
            f"""
            SELECT i.item_text, COUNT(*) AS cnt
            FROM inspection_run_items i
            JOIN inspection_runs r ON r.id = i.run_id
            WHERE {where_sql} AND upper(i.result)='NONCOMPLIANT'
            GROUP BY i.item_text
            ORDER BY cnt DESC, i.item_text ASC
            LIMIT 10
            """,
            tuple(params),
        ).fetchall()
        return {"ok": True, "site_code": scoped_site, "totals": dict(totals) if totals else {}, "top_noncompliant_items": [dict(x) for x in top_rows]}
    finally:
        con.close()
