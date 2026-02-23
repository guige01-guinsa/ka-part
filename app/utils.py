from __future__ import annotations

import datetime as dt
import logging
import re
from pathlib import Path
from typing import Any, Dict, List

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from .schema_defs import SCHEMA_DEFS, SCHEMA_TAB_ORDER

LOG = logging.getLogger(__name__)
_WEASYPRINT_HTML_CLASS: Any | None = None
_WEASYPRINT_AVAILABLE: bool | None = None
PDF_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates" / "pdf"
DEFAULT_PDF_PROFILE_ID = "substation_daily_a4"
DEFAULT_PDF_TEMPLATE_NAME = "substation_daily_a4.html"
PDF_PROFILE_DEFS: Dict[str, Dict[str, str]] = {
    "substation_daily_a4": {
        "template_name": "substation_daily_a4.html",
        "context_builder": "substation",
    },
    "substation_daily_ami4_a4": {
        "template_name": "substation_daily_ami4_a4.html",
        "context_builder": "substation",
    },
    "substation_daily_generic_a4": {
        "template_name": "substation_daily_generic_a4.html",
        "context_builder": "generic",
    },
}
_SAFE_PDF_TOKEN_RE = re.compile(r"[^A-Za-z0-9._-]+")
_PDF_PAGE_BLOCK_RE = re.compile(r"@page(?:\s+[^{]+)?\s*\{[^{}]*\}", re.IGNORECASE | re.DOTALL)
_PDF_PAGE_MARGIN_RE = re.compile(r"margin\s*:\s*[^;{}]+;?", re.IGNORECASE)
_HTML_HEAD_CLOSE_RE = re.compile(r"</head\s*>", re.IGNORECASE)
PDF_PAGE_MARGIN_TOP_MM = 20.0
PDF_PAGE_MARGIN_RIGHT_MM = 8.0
PDF_PAGE_MARGIN_BOTTOM_MM = 15.0
PDF_PAGE_MARGIN_LEFT_MM = 8.0
PDF_PAGE_HEIGHT_MM = 297.0

def today_ymd() -> str:
    return dt.date.today().isoformat()

def safe_ymd(s: str) -> str:
    if not s:
        return today_ymd()
    try:
        return dt.date.fromisoformat(str(s)[:10]).isoformat()
    except Exception:
        return today_ymd()

def _flatten_tabs(tabs: Dict[str, Dict[str, str]]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for tab_key, fields in (tabs or {}).items():
        if not isinstance(fields, dict):
            continue
        for k, v in fields.items():
            out[f"{tab_key}.{k}"] = str(v)
    return out


def _ordered_export_keys(rows: List[Dict[str, Any]], schema_defs: Dict[str, Dict[str, Any]] | None = None) -> List[str]:
    """Keep export column order stable and schema-driven."""
    source = schema_defs if isinstance(schema_defs, dict) else SCHEMA_DEFS
    present = set()
    for r in rows:
        flat = _flatten_tabs(r.get("tabs") or {})
        present.update(flat.keys())

    ordered: List[str] = []
    tab_order = [t for t in SCHEMA_TAB_ORDER if t in source] + [t for t in source.keys() if t not in SCHEMA_TAB_ORDER]
    for tab_key in tab_order:
        tab = source.get(tab_key) or {}
        for f in tab.get("fields") or []:
            key = f"{tab_key}.{f.get('k')}"
            if key in present:
                ordered.append(key)
                present.remove(key)

    ordered.extend(sorted(present))
    return ordered

def build_excel(
    site_name: str,
    date_from: str,
    date_to: str,
    rows: List[Dict[str, Any]],
    *,
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "entries"

    export_keys = _ordered_export_keys(rows, schema_defs=schema_defs)
    include_work_type = any(str(r.get("work_type") or "").strip() for r in rows)
    flattened = [
        (
            r.get("entry_date", ""),
            str(r.get("work_type") or "").strip(),
            _flatten_tabs(r.get("tabs") or {}),
        )
        for r in rows
    ]
    headers = ["entry_date"] + (["work_type"] if include_work_type else []) + export_keys
    ws.append(headers)

    for entry_date, work_type, flat in flattened:
        lead = [entry_date] + ([work_type] if include_work_type else [])
        ws.append(lead + [flat.get(k, "") for k in export_keys])

    # Automatic width optimization by real content length.
    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(48, max_len + 2))
    ws.freeze_panes = "A2"

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def _fmt_value(value: Any, default: str = "-") -> str:
    text = str(value or "").strip()
    return text if text else default


def _tab_value(
    tabs: Dict[str, Dict[str, Any]],
    tab_key: str,
    field_key: str,
    *,
    default: str = "-",
) -> str:
    tab = tabs.get(tab_key) if isinstance(tabs, dict) else {}
    if not isinstance(tab, dict):
        return default
    return _fmt_value(tab.get(field_key), default=default)


def _clean_pdf_profile_id(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    clean = _SAFE_PDF_TOKEN_RE.sub("", raw)
    return clean[:80]


def _clean_pdf_template_name(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    name = raw.replace("\\", "/").split("/")[-1].strip()
    if not name or name in {".", ".."}:
        return ""
    if not name.lower().endswith(".html"):
        return ""
    clean = _SAFE_PDF_TOKEN_RE.sub("", name)
    if not clean.lower().endswith(".html"):
        return ""
    return clean[:120]


def _pdf_template_exists(template_name: str) -> bool:
    clean = _clean_pdf_template_name(template_name)
    if not clean:
        return False
    return (PDF_TEMPLATE_DIR / clean).is_file()


def _resolve_pdf_render_plan(site_env_config: Dict[str, Any] | None = None) -> Dict[str, Any]:
    cfg = site_env_config if isinstance(site_env_config, dict) else {}
    report_cfg = cfg.get("report") if isinstance(cfg.get("report"), dict) else {}
    requested_profile_id = _clean_pdf_profile_id(report_cfg.get("pdf_profile_id"))
    requested_template_name = _clean_pdf_template_name(report_cfg.get("pdf_template_name"))
    locked_profile_id = _clean_pdf_profile_id(report_cfg.get("locked_profile_id"))

    effective_profile_id = requested_profile_id
    locked_profile_applied = False
    if locked_profile_id:
        if locked_profile_id in PDF_PROFILE_DEFS:
            if requested_profile_id and requested_profile_id != locked_profile_id:
                LOG.info(
                    "PDF profile is locked. requested='%s' -> locked='%s'.",
                    requested_profile_id,
                    locked_profile_id,
                )
            effective_profile_id = locked_profile_id
            locked_profile_applied = True
            # Locked profile ignores per-site custom template override.
            requested_template_name = ""
        else:
            LOG.warning("Unknown PDF locked_profile_id '%s'; ignore lock.", locked_profile_id)

    profile = PDF_PROFILE_DEFS.get(effective_profile_id) if effective_profile_id else None
    if profile is None:
        if effective_profile_id:
            kind = "locked_profile_id" if locked_profile_applied else "pdf_profile_id"
            LOG.warning("Unknown PDF %s '%s'; fallback to '%s'.", kind, effective_profile_id, DEFAULT_PDF_PROFILE_ID)
        profile_id = DEFAULT_PDF_PROFILE_ID
        profile = PDF_PROFILE_DEFS.get(profile_id) or {}
    else:
        profile_id = effective_profile_id

    profile_template = _clean_pdf_template_name(profile.get("template_name"))
    template_name = requested_template_name or profile_template or DEFAULT_PDF_TEMPLATE_NAME
    if not _pdf_template_exists(template_name):
        if requested_template_name:
            LOG.warning("Configured PDF template '%s' not found; fallback to profile template.", requested_template_name)
        template_name = profile_template or DEFAULT_PDF_TEMPLATE_NAME
    if not _pdf_template_exists(template_name):
        LOG.warning("Profile template '%s' not found; fallback to '%s'.", template_name, DEFAULT_PDF_TEMPLATE_NAME)
        template_name = DEFAULT_PDF_TEMPLATE_NAME
        profile_id = DEFAULT_PDF_PROFILE_ID
        profile = PDF_PROFILE_DEFS.get(profile_id) or {}

    context_builder = str(profile.get("context_builder") or "substation").strip().lower()
    if context_builder not in {"substation", "generic"}:
        context_builder = "substation"

    return {
        "profile_id": profile_id,
        "template_name": template_name,
        "context_builder": context_builder,
        "requested_profile_id": requested_profile_id,
        "requested_template_name": requested_template_name,
        "locked_profile_id": locked_profile_id,
    }


def _format_mm(value_mm: float) -> str:
    text = f"{float(value_mm):.2f}".rstrip("0").rstrip(".")
    return text or "0"


def _fixed_page_margin_css() -> str:
    return (
        f"{_format_mm(PDF_PAGE_MARGIN_TOP_MM)}mm "
        f"{_format_mm(PDF_PAGE_MARGIN_RIGHT_MM)}mm "
        f"{_format_mm(PDF_PAGE_MARGIN_BOTTOM_MM)}mm "
        f"{_format_mm(PDF_PAGE_MARGIN_LEFT_MM)}mm"
    )


def _fixed_content_height_mm() -> float:
    return max(0.0, PDF_PAGE_HEIGHT_MM - PDF_PAGE_MARGIN_TOP_MM - PDF_PAGE_MARGIN_BOTTOM_MM)


def _apply_pdf_page_margin(html: str) -> str:
    source = str(html or "")
    if not source:
        return source
    margin_decl = f"margin: {_fixed_page_margin_css()};"

    def _replace_page_block(match: re.Match[str]) -> str:
        block = match.group(0)
        if _PDF_PAGE_MARGIN_RE.search(block):
            return _PDF_PAGE_MARGIN_RE.sub(margin_decl, block, count=1)
        pos = block.rfind("}")
        if pos < 0:
            return block
        return f"{block[:pos]} {margin_decl} {block[pos:]}"

    updated = _PDF_PAGE_BLOCK_RE.sub(_replace_page_block, source)
    if updated == source:
        updated = f"<style>@page {{ size: A4; {margin_decl} }}</style>\n{updated}"

    fit_inject = (
        "<style>"
        f".sheet{{min-height:{_format_mm(_fixed_content_height_mm())}mm;box-sizing:border-box;}}"
        "</style>"
    )
    head_match = _HTML_HEAD_CLOSE_RE.search(updated)
    if head_match:
        idx = head_match.start()
        return f"{updated[:idx]}{fit_inject}\n{updated[idx:]}"
    return f"{fit_inject}\n{updated}"


def _keep_pdf_first_page(pdf_bytes: bytes) -> bytes:
    raw = bytes(pdf_bytes or b"")
    if not raw:
        return raw
    try:
        from io import BytesIO
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return raw
    try:
        reader = PdfReader(BytesIO(raw))
        if len(reader.pages) <= 1:
            return raw
        writer = PdfWriter()
        writer.add_page(reader.pages[0])
        out = BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        return raw


def _date_label_ko(ymd: str) -> str:
    try:
        d = dt.date.fromisoformat(safe_ymd(ymd))
    except Exception:
        d = dt.date.today()
    weeks = ["월", "화", "수", "목", "금", "토", "일"]
    return f"{d.year}년 {d.month}월 {d.day}일 {weeks[d.weekday()]}요일"


def _build_pdf_context(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, Any]],
    *,
    worker_name: str = "",
) -> Dict[str, Any]:
    home = tabs.get("home") if isinstance(tabs, dict) else {}
    if not isinstance(home, dict):
        home = {}

    complex_code = _fmt_value(home.get("complex_code"), default="")
    if not complex_code:
        complex_code = _fmt_value(site_name, default="")

    lv_rows = ("L1", "L2", "L3")
    lv_cols = (
        ("V", "_V"),
        ("A", "_A"),
        ("KW", "_KW"),
    )

    def _to_float(value: Any) -> float | None:
        txt = str(value or "").strip()
        if not txt:
            return None
        try:
            return float(txt.replace(",", ""))
        except Exception:
            return None

    def _fmt_num(value: float | None) -> str:
        if value is None:
            return "-"
        if abs(value - round(value)) < 0.0001:
            return str(int(round(value)))
        return f"{value:.2f}".rstrip("0").rstrip(".")

    def _tab_avg(tab_key: str, field_keys: List[str]) -> str:
        tab = tabs.get(tab_key) if isinstance(tabs, dict) else {}
        if not isinstance(tab, dict):
            return "-"
        nums: List[float] = []
        for key in field_keys:
            n = _to_float(tab.get(key))
            if n is not None:
                nums.append(n)
        if not nums:
            return "-"
        return _fmt_num(sum(nums) / len(nums))

    def lv_block(prefix: str) -> List[Dict[str, str]]:
        rows: List[Dict[str, str]] = []
        for row_name in lv_rows:
            item: Dict[str, str] = {"phase": row_name}
            for col_name, suffix in lv_cols:
                key = f"{prefix}_{row_name}{suffix}"
                item[col_name.lower()] = _tab_value(tabs, "tr1" if prefix == "lv1" else "tr2", key)
            rows.append(item)
        return rows

    meter_main = _tab_value(tabs, "meter", "main_kwh")
    meter_industry = _tab_value(tabs, "meter", "industry_kwh")
    meter_street = _tab_value(tabs, "meter", "street_kwh")
    main_vcb_r = _tab_value(tabs, "main_vcb", "main_vcb_l1_a")
    main_vcb_s = _tab_value(tabs, "main_vcb", "main_vcb_l2_a")
    main_vcb_t = _tab_value(tabs, "main_vcb", "main_vcb_l3_a")

    return {
        "title": "수배전반(검침)점검일지",
        "site_name": _fmt_value(site_name, default="-"),
        "site_office_label": f"{_fmt_value(site_name, default='-')} 관리사무소",
        "site_code": _fmt_value(complex_code, default="-"),
        "entry_date": safe_ymd(date),
        "entry_date_label": _date_label_ko(date),
        "worker_name": _fmt_value(worker_name, default="-"),
        "weather": _tab_value(tabs, "home", "weather", default=""),
        "approval_staff": _tab_value(tabs, "home", "approval_staff", default=""),
        "approval_manager": _tab_value(tabs, "home", "approval_manager", default=""),
        "approval_deputy": _tab_value(tabs, "home", "approval_deputy", default=""),
        "approval_chief": _tab_value(tabs, "home", "approval_chief", default=""),
        "work_type": _fmt_value(home.get("work_type"), default="일일"),
        "important_work": _fmt_value(home.get("important_work"), default="-"),
        "note": _fmt_value(home.get("note"), default="-"),
        "inspection_time": "07:30",
        "aiss_kv": _tab_value(tabs, "main_vcb", "main_vcb_kv"),
        "aiss_r_a": _tab_value(tabs, "meter", "AISS_L1_A"),
        "aiss_s_a": _tab_value(tabs, "meter", "AISS_L2_A"),
        "aiss_t_a": _tab_value(tabs, "meter", "AISS_L3_A"),
        # Current schema has no neutral current key; keep explicit placeholder.
        "aiss_n_a": "0",
        "tr1_temp": _tab_value(tabs, "temperature", "temperature_tr1"),
        "tr2_temp": _tab_value(tabs, "temperature", "temperature_tr2"),
        "tr3_temp": _tab_value(tabs, "temperature", "temperature_tr3"),
        "lv1_rows": lv_block("lv1"),
        "lv2_rows": lv_block("lv2"),
        "main_vcb_r": main_vcb_r,
        "main_vcb_s": main_vcb_s,
        "main_vcb_t": main_vcb_t,
        "main_vcb_a_avg": _tab_avg("main_vcb", ["main_vcb_l1_a", "main_vcb_l2_a", "main_vcb_l3_a"]),
        "main_vcb_kw": _tab_value(tabs, "main_vcb", "main_vcb_kw"),
        "main_vcb_pf": _tab_value(tabs, "main_vcb", "main_vcb_pf"),
        "lv1_v_avg": _tab_avg("tr1", ["lv1_L1_V", "lv1_L2_V", "lv1_L3_V"]),
        "lv1_a_avg": _tab_avg("tr1", ["lv1_L1_A", "lv1_L2_A", "lv1_L3_A"]),
        "lv1_kw_avg": _tab_avg("tr1", ["lv1_L1_KW", "lv1_L2_KW", "lv1_L3_KW"]),
        "lv3_v_avg": _tab_avg("tr3", ["lv3_L1_V", "lv3_L2_V", "lv3_L3_V"]),
        "lv3_a_avg": _tab_avg("tr3", ["lv3_L1_A", "lv3_L2_A", "lv3_L3_A"]),
        "lv3_kw_avg": _tab_avg("tr3", ["lv3_L1_KW", "lv3_L2_KW", "lv3_L3_KW"]),
        "lv5_v_avg": _tab_avg("tr5", ["lv5_L1_V", "lv5_L2_V", "lv5_L3_V"]),
        "lv5_a_avg": _tab_avg("tr5", ["lv5_L1_A", "lv5_L2_A", "lv5_L3_A"]),
        "lv5_kw_avg": _tab_avg("tr5", ["lv5_L1_KW", "lv5_L2_KW", "lv5_L3_KW"]),
        "dc_panel_v": _tab_value(tabs, "dc_panel", "dc_panel_v"),
        "dc_panel_a": _tab_value(tabs, "dc_panel", "dc_panel_a"),
        "meter_main": meter_main,
        "meter_industry": meter_industry,
        "meter_street": meter_street,
        "meter_rows": [
            {"name": "메인(*720)④", "today": meter_main, "prev": "#N/A", "daily": "", "monthly": ""},
            {"name": "산업용③", "today": meter_industry, "prev": "#N/A", "daily": "", "monthly": ""},
            {"name": "가로등⑥", "today": meter_street, "prev": "#N/A", "daily": "", "monthly": ""},
        ],
        "tank_apartment": _tab_value(tabs, "facility_check", "tank_level_1"),
        "tank_officetel": _tab_value(tabs, "facility_check", "tank_level_2"),
        "hydrant_pressure": _tab_value(tabs, "facility_check", "hydrant_pressure"),
        "sp_pump_pressure": _tab_value(tabs, "facility_check", "sp_pump_pressure"),
        "high_pressure": _tab_value(tabs, "facility_check", "high_pressure"),
        "low_pressure": _tab_value(tabs, "facility_check", "low_pressure"),
        "office_pressure": _tab_value(tabs, "facility_check", "office_pressure"),
        "shop_pressure": _tab_value(tabs, "facility_check", "shop_pressure"),
    }


def _build_pdf_context_generic(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, Any]],
    *,
    worker_name: str = "",
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
) -> Dict[str, Any]:
    source = schema_defs if isinstance(schema_defs, dict) else SCHEMA_DEFS
    home = tabs.get("home") if isinstance(tabs, dict) else {}
    if not isinstance(home, dict):
        home = {}

    site_code = _fmt_value(home.get("complex_code"), default="")
    if not site_code:
        site_code = _fmt_value(site_name, default="")

    order_map = {k: i for i, k in enumerate(SCHEMA_TAB_ORDER)}
    sections: List[Dict[str, Any]] = []
    for tab_key in sorted((tabs or {}).keys(), key=lambda x: (order_map.get(x, 999), x)):
        fields = tabs.get(tab_key)
        if not isinstance(fields, dict):
            continue
        tab_def = source.get(tab_key) or {}
        title = str(tab_def.get("title") or tab_key)
        label_map = {
            str(f.get("k")): str(f.get("label") or f.get("k") or "")
            for f in (tab_def.get("fields") or [])
            if isinstance(f, dict) and f.get("k")
        }
        rows = [
            {
                "key": str(k),
                "label": label_map.get(str(k), str(k)),
                "value": _fmt_value(v, default="-"),
            }
            for k, v in sorted(fields.items(), key=lambda kv: str(kv[0]))
        ]
        if rows:
            sections.append({"tab_key": tab_key, "title": title, "rows": rows})

    if not sections:
        sections.append(
            {
                "tab_key": "empty",
                "title": "입력 항목",
                "rows": [{"key": "-", "label": "안내", "value": "저장된 점검 데이터가 없습니다."}],
            }
        )

    return {
        "title": "시설 점검일지",
        "site_name": _fmt_value(site_name, default="-"),
        "site_office_label": f"{_fmt_value(site_name, default='-')} 관리사무소",
        "site_code": _fmt_value(site_code, default="-"),
        "entry_date": safe_ymd(date),
        "entry_date_label": _date_label_ko(date),
        "worker_name": _fmt_value(worker_name, default="-"),
        "work_type": _fmt_value(home.get("work_type"), default="일일"),
        "note": _fmt_value(home.get("note"), default="-"),
        "sections": sections,
    }


def _render_pdf_html_template(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, Any]],
    *,
    worker_name: str = "",
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
    site_env_config: Dict[str, Any] | None = None,
) -> tuple[str, Path, str]:
    template_dir = PDF_TEMPLATE_DIR
    plan = _resolve_pdf_render_plan(site_env_config=site_env_config)
    context_builder = plan.get("context_builder") or "substation"
    page_margin_css = _fixed_page_margin_css()

    if context_builder == "generic":
        context = _build_pdf_context_generic(
            site_name,
            date,
            tabs,
            worker_name=worker_name,
            schema_defs=schema_defs,
        )
    else:
        context = _build_pdf_context(site_name, date, tabs, worker_name=worker_name)
    context["profile_id"] = plan.get("profile_id") or DEFAULT_PDF_PROFILE_ID
    context["template_name"] = plan.get("template_name") or DEFAULT_PDF_TEMPLATE_NAME
    context["page_margin_css"] = page_margin_css

    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template_name = context["template_name"]
    try:
        tpl = env.get_template(template_name)
    except Exception:
        template_name = DEFAULT_PDF_TEMPLATE_NAME
        context = _build_pdf_context(site_name, date, tabs, worker_name=worker_name)
        context["template_name"] = template_name
        context["profile_id"] = DEFAULT_PDF_PROFILE_ID
        context["page_margin_css"] = page_margin_css
        LOG.warning("PDF template load failed. fallback template='%s'.", template_name)
        tpl = env.get_template(template_name)
    html = _apply_pdf_page_margin(tpl.render(**context))
    return html, template_dir, template_name


def _get_weasyprint_html_class():
    global _WEASYPRINT_HTML_CLASS, _WEASYPRINT_AVAILABLE
    if _WEASYPRINT_AVAILABLE is True and _WEASYPRINT_HTML_CLASS is not None:
        return _WEASYPRINT_HTML_CLASS
    if _WEASYPRINT_AVAILABLE is False:
        return None
    try:
        from weasyprint import HTML as html_class
    except Exception as e:
        _WEASYPRINT_AVAILABLE = False
        LOG.warning("WeasyPrint unavailable, using xhtml2pdf fallback: %s", e)
        return None
    _WEASYPRINT_HTML_CLASS = html_class
    _WEASYPRINT_AVAILABLE = True
    return _WEASYPRINT_HTML_CLASS


def _build_pdf_html_weasyprint(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, Any]],
    *,
    worker_name: str = "",
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
    site_env_config: Dict[str, Any] | None = None,
) -> bytes:
    html, template_dir, _template_name = _render_pdf_html_template(
        site_name,
        date,
        tabs,
        worker_name=worker_name,
        schema_defs=schema_defs,
        site_env_config=site_env_config,
    )
    html_class = _get_weasyprint_html_class()
    if html_class is None:
        raise RuntimeError("WeasyPrint is unavailable")
    pdf = html_class(string=html, base_url=str(template_dir)).write_pdf()
    return bytes(pdf)


def _build_pdf_html_xhtml2pdf(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, Any]],
    *,
    worker_name: str = "",
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
    site_env_config: Dict[str, Any] | None = None,
) -> bytes:
    from io import BytesIO
    from xhtml2pdf import pisa

    html, template_dir, template_name = _render_pdf_html_template(
        site_name,
        date,
        tabs,
        worker_name=worker_name,
        schema_defs=schema_defs,
        site_env_config=site_env_config,
    )
    template_path = (template_dir / template_name).resolve()
    out = BytesIO()
    result = pisa.CreatePDF(src=html, dest=out, encoding="utf-8", path=str(template_path))
    if getattr(result, "err", 0):
        raise RuntimeError(f"xhtml2pdf render failed: err={result.err}")
    return out.getvalue()


def _build_pdf_legacy(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, str]],
    *,
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
    render_plan: Dict[str, Any] | None = None,
) -> bytes:
    from io import BytesIO
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    width, height = A4
    plan = render_plan if isinstance(render_plan, dict) else {}
    context_builder = str(plan.get("context_builder") or "substation").strip().lower()
    profile_id = str(plan.get("profile_id") or DEFAULT_PDF_PROFILE_ID).strip() or DEFAULT_PDF_PROFILE_ID
    template_name = str(plan.get("template_name") or DEFAULT_PDF_TEMPLATE_NAME).strip() or DEFAULT_PDF_TEMPLATE_NAME
    title_text = "수변전실 점검일지" if context_builder == "substation" else "시설 점검일지"

    y = height - 40
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, y, title_text)
    y -= 14
    c.setFont("Helvetica", 8)
    c.drawString(40, y, f"profile: {profile_id} / template: {template_name}")
    y -= 22
    c.setFont("Helvetica", 11)
    c.drawString(40, y, f"단지: {site_name}    날짜: {date}")
    y -= 18
    c.line(40, y, width-40, y)
    y -= 18

    c.setFont("Helvetica", 10)
    source = schema_defs if isinstance(schema_defs, dict) else SCHEMA_DEFS
    order_map = {k: i for i, k in enumerate(SCHEMA_TAB_ORDER)}
    for tab_key in sorted((tabs or {}).keys(), key=lambda x: (order_map.get(x, 999), x)):
        c.setFont("Helvetica-Bold", 11)
        title = (source.get(tab_key) or {}).get("title") or tab_key
        c.drawString(40, y, f"[{title}]")
        y -= 14
        c.setFont("Helvetica", 10)
        fields = tabs.get(tab_key) or {}
        label_map = {
            str(f.get("k")): str(f.get("label"))
            for f in ((source.get(tab_key) or {}).get("fields") or [])
            if isinstance(f, dict) and f.get("k")
        }
        for k in sorted(fields.keys()):
            v = str(fields.get(k, ""))
            label = label_map.get(k, k)
            line = f"- {label}({k}): {v}"
            c.drawString(50, y, line[:120])
            y -= 12
            if y < 60:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 10)
        y -= 6
        if y < 60:
            c.showPage()
            y = height - 40
            c.setFont("Helvetica", 10)

    c.showPage()
    c.save()
    return bio.getvalue()


def build_pdf(
    site_name: str,
    date: str,
    tabs: Dict[str, Dict[str, str]],
    *,
    worker_name: str = "",
    schema_defs: Dict[str, Dict[str, Any]] | None = None,
    site_env_config: Dict[str, Any] | None = None,
) -> bytes:
    render_plan = _resolve_pdf_render_plan(site_env_config=site_env_config)
    if _get_weasyprint_html_class() is not None:
        try:
            pdf = _build_pdf_html_weasyprint(
                site_name,
                date,
                tabs or {},
                worker_name=worker_name,
                schema_defs=schema_defs,
                site_env_config=site_env_config,
            )
            return _keep_pdf_first_page(pdf)
        except Exception as e:
            LOG.warning("WeasyPrint render failed, trying xhtml2pdf fallback: %s", e)
    try:
        pdf = _build_pdf_html_xhtml2pdf(
            site_name,
            date,
            tabs or {},
            worker_name=worker_name,
            schema_defs=schema_defs,
            site_env_config=site_env_config,
        )
        return _keep_pdf_first_page(pdf)
    except Exception as e:
        LOG.warning("xhtml2pdf render failed, fallback to legacy reportlab: %s", e)
    return _keep_pdf_first_page(
        _build_pdf_legacy(
            site_name,
            date,
            tabs,
            schema_defs=schema_defs,
            render_plan=render_plan,
        )
    )
